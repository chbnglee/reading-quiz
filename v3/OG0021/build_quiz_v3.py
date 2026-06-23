# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "OG0021_ReadingQuiz.xlsx"

C_HEADER = "1F4E79"
C_SUB = "2E75B6"
C_ACCENT = "BDD7EE"
C_YELLOW = "FFF2CC"
C_GREEN = "E2EFDA"
C_OK = "70AD47"
C_PART = "FFD966"
C_LOW = "FFEECC"
C_BAD = "FFCCCC"


def style_cell(c, fill=None, bold=False, color="000000", size=9, align="left"):
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    return c


def header(ws, row, col, value, fill=C_SUB, size=10):
    return style_cell(ws.cell(row, col, value), fill=fill, bold=True, color="FFFFFF", size=size, align="center")


def merge_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    style_cell(c, C_HEADER, True, "FFFFFF", 12, "center")
    c.value = text


def set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def option_fill(score):
    if score == 100:
        return C_OK
    if score >= 30:
        return C_PART
    if score > 0:
        return C_LOW
    return C_BAD


def make_option_sheet(title, sheet, question, options, lrs, resource="-"):
    ws = wb.create_sheet(sheet)
    set_widths(ws, [8, 36, 12, 20, 20, 24, 28, 28, 22])
    merge_title(ws, "A1:I1", title)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:I2")
    style_cell(ws["A2"], C_YELLOW, True, "000000", 9).value = "Question: " + question
    if resource != "-":
        ws.merge_cells("A3:I3")
        style_cell(ws["A3"], C_ACCENT, False, "000000", 9).value = "Resource: " + resource
    hdrs = [
        "Option",
        "Option Text",
        "Is Correct?",
        "Score",
        "SG Element",
        "Distractor Type",
        "Distractor Rationale",
        "Student Weakness Signal",
        "LRS sg_element",
    ]
    start = 5 if resource != "-" else 4
    for i, h in enumerate(hdrs, 1):
        header(ws, start, i, h)
    for r, row in enumerate(options, start + 1):
        fill = option_fill(row[3])
        for c, v in enumerate(row, 1):
            bg = fill if c in (1, 3, 4) else None
            fg = "FFFFFF" if bg == C_OK and c in (1, 3, 4) else "000000"
            style_cell(ws.cell(r, c, v), bg, c in (1, 3, 4) and row[3] == 100, fg)
        ws.row_dimensions[r].height = 36
    foot = start + len(options) + 2
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=9)
    style_cell(ws.cell(foot, 1), C_GREEN, True, "000000", 9).value = (
        f'LRS xAPI: verb="answered" | object="quiz_OG0021_{sheet.lower()}" '
        f'| result.sg_element="{lrs}" | result.score_raw=<pts>'
    )


def make_sequence_sheet(title, sheet, question, rows, lrs):
    ws = wb.create_sheet(sheet)
    set_widths(ws, [18, 20, 14, 14, 24, 36, 24, 24])
    merge_title(ws, "A1:H1", title)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:H2")
    style_cell(ws["A2"], C_YELLOW, True).value = "Question: " + question
    for i, h in enumerate(
        ["Scene", "Resource", "Correct Pos", "Scene Weight", "SG Role", "Weight Rationale", "Error Tag", "LRS"],
        1,
    ):
        header(ws, 4, i, h)
    for r, row in enumerate(rows, 5):
        row_fill = C_YELLOW if float(row[3]) >= 2.5 else C_GREEN
        for c, v in enumerate(row, 1):
            style_cell(ws.cell(r, c, v), row_fill if c in (1, 2, 3, 4, 5) else None)
        ws.row_dimensions[r].height = 30
    total_row = 5 + len(rows)
    for c, v in [(1, "TOTAL"), (4, f"=SUM(D5:D{total_row - 1})")]:
        style_cell(ws.cell(total_row, c, v), C_ACCENT, True, "000000", 9, "center")

    matrix_title = total_row + 2
    ws.merge_cells(start_row=matrix_title, start_column=1, end_row=matrix_title, end_column=8)
    header(ws, matrix_title, 1, "SECTION B - Score Matrix: Points per (Scene, Submitted Position)")
    matrix_header = matrix_title + 1
    header(ws, matrix_header, 1, "Scene \\ Submitted Pos", C_ACCENT)
    for pos in range(1, 6):
        header(ws, matrix_header, pos + 1, f"Pos {pos}", C_SUB)
    for idx, row in enumerate(rows):
        source_row = 5 + idx
        out_row = matrix_header + 1 + idx
        correct_pos = int(row[2])
        style_cell(ws.cell(out_row, 1, row[0]), C_ACCENT, True, "000000", 9, "center")
        for pos in range(1, 6):
            distance = abs(pos - correct_pos)
            fill = C_OK if distance == 0 else C_PART if distance == 1 else C_BAD
            font_color = "FFFFFF" if distance == 0 else "000000"
            formula = f"=ROUND($D${source_row}/SUM($D$5:$D${total_row - 1})*100*MAX(0,1-ABS({pos}-$C${source_row})*0.5),1)"
            style_cell(ws.cell(out_row, pos + 1, formula), fill, distance == 0, font_color, 9, "center")

    example_title = matrix_header + len(rows) + 3
    ws.merge_cells(start_row=example_title, start_column=1, end_row=example_title, end_column=8)
    header(ws, example_title, 1, "SECTION C - Example Answer Scores")
    for i, h in enumerate(["Example Student Answer", "Submitted Order", "Score", "Interpretation"], 1):
        header(ws, example_title + 1, i, h, C_ACCENT)
    examples = [
        ["Perfect", "SC01-SC02-SC03-SC06-SC09", 100, "Full story arc understood."],
        ["Middle scenes swapped", "SC01-SC02-SC06-SC03-SC09", 84, "Core anchors are intact; attempt/reaction order is unstable."],
        ["Problem/result misplaced", "SC02-SC01-SC03-SC09-SC06", 47, "Key consequence positions are confused."],
    ]
    for r, row in enumerate(examples, example_title + 2):
        for c, v in enumerate(row, 1):
            style_cell(ws.cell(r, c, v), option_fill(v) if c == 3 else None)

    foot = example_title + len(examples) + 3
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=8)
    style_cell(ws.cell(foot, 1), C_GREEN, True).value = (
        f'LRS xAPI: verb="answered" | object="quiz_OG0021_{sheet.lower()}" '
        f'| result.sg_element="{lrs}" | result.score_raw=<pts>'
    )


wb = Workbook()

# QUIZ_LIST
ws = wb.active
ws.title = "QUIZ_LIST"
set_widths(ws, [8, 22, 22, 34, 26, 28, 12, 22, 24, 20, 42])
merge_title(ws, "A1:K1", "OG0021 Reading Quiz v3 - Story Grammar")
ws.row_dimensions[1].height = 28
cols = [
    "Q_ID",
    "Quiz Type",
    "Primary Story Grammar",
    "Question (EN)",
    "Resource",
    "Correct Answer",
    "Max Score",
    "LRS sg_element",
    "Scoring Mode",
    "Sheet Ref",
    "Hint (A1 EN)",
]
for i, h in enumerate(cols, 1):
    header(ws, 2, i, h)
ws.row_dimensions[2].height = 30
rows = [
    ["Q01", "Story Scene Sequence", "Consequence", "Put the story scenes in order.", "SC01/02/03/06/09 images", "SC01 -> SC02 -> SC03 -> SC06 -> SC09", 100, "consequence", "Weighted Position", "Q01_CONSEQUENCE", "Milo loses his color. Put the scenes from first to last."],
    ["Q02", "Setting Slot Drag", "Setting", "Look at the first scene. Fill in the boxes.", "SC01 image + word cards", "who=chameleon / where=forest / at_first=loves changing colors", 100, "setting", "Weighted Slot Match", "Q02_SETTING", "Look at the first picture. Who is there? Where is he?"],
    ["Q03", "Listening Scene Match", "Initiating Event", "Listen. Which scene starts the problem?", "Audio + SC02/03/06/09", "Option A", 100, "initiating_event", "Weighted MCQ", "Q03_INIT_EVENT", "Listen for the problem. What starts the story?"],
    ["Q04", "Scene-Anchored Unscramble", "Attempt", "Put the story words in order.", "SC03 image", "Milo walks into the forest.", 100, "attempt", "Weighted Unscramble", "Q04_ATTEMPT", "Milo goes to find his color. Build the same sentence."],
    ["Q05", "Feeling Match", "Reaction", "How does Milo feel here?", "SC06 image", "Option B", 100, "reaction", "Weighted MCQ", "Q05_REACTION", "Look at Milo's face. How does he feel?"],
    ["Q06", "Internal Response MCQ", "Internal Response", "What is Milo thinking?", "SC06 image", "Option A", 100, "internal_response", "Weighted MCQ", "Q06_INTERNAL", "Think about Milo's heart. What does he learn?"],
]
for r, row in enumerate(rows, 3):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v))
    ws.row_dimensions[r].height = 28
ws.merge_cells("A10:K10")
style_cell(ws["A10"], C_GREEN, False, "000000", 9).value = (
    "v3 keeps one primary Story Grammar score per Q01-Q06 and removes the extra whole-story item from quiz, report, and overall scoring. "
    "Bookey hints are short A1-level English supports shown inside the quiz UI."
)

# Q01 Consequence
make_sequence_sheet(
    "Q01 - Story Scene Sequence | Story Grammar: Consequence",
    "Q01_CONSEQUENCE",
    "Put the story scenes in order.",
    [
        ["SC01", "OG0021_SC01_I", 1, 1.5, "Opening state", "Shows Milo before the problem begins; useful context but lower weight than problem/result.", "missed_opening_state", "consequence"],
        ["SC02", "OG0021_SC02_I", 2, 2.5, "Problem begins", "Milo wakes up gray; this key event drives the rest of the story.", "missed_problem_start", "initiating_event"],
        ["SC03", "OG0021_SC03_I", 3, 1.5, "Attempt begins", "Milo starts searching in the forest; a middle bridge scene.", "attempt_position_shift", "attempt"],
        ["SC06", "OG0021_SC06_I", 4, 1.5, "Low point / reaction", "Milo cries by the pond; shows the consequence before recovery.", "reaction_position_shift", "reaction"],
        ["SC09", "OG0021_SC09_I", 5, 2.5, "Final result", "Milo's colors come back; final consequence gets high weight.", "missed_final_result", "consequence"],
    ],
    "consequence",
)

# Q02 Setting
ws = wb.create_sheet("Q02_SETTING")
set_widths(ws, [12, 18, 18, 20, 16, 14, 20, 34, 28])
merge_title(ws, "A1:I1", "Q02 - Setting Slot Drag | Story Grammar: Setting")
ws.merge_cells("A2:I2")
style_cell(ws["A2"], C_YELLOW, True).value = "Question: Look at the first scene. Fill in the boxes."
ws.merge_cells("A3:I3")
style_cell(ws["A3"], C_ACCENT, False).value = "Scene resource: OG0021_SC01_I.png"
for i, h in enumerate(["Card Key", "Text", "Slot", "Resource", "Correct?", "Slot Weight", "Credit Rule", "Distractor Rationale", "LRS"], 1):
    header(ws, 4, i, h)
setting_rows = [
    ["pond", "pond", "Where?", "OG0021_SC06_I", "NO", 2.0, "35% slot credit if placed in Where", "Uses a later important place as the opening place.", "setting"],
    ["chameleon", "chameleon", "Who?", "OG0021_SC01_I", "YES", 2.5, "100% if placed in Who", "Correct main character at the beginning.", "setting"],
    ["loses_color", "loses his color", "At first...", "OG0021_SC02_I", "NO", 1.5, "35% slot credit if placed in At first", "Uses the problem event as the opening state.", "setting"],
    ["forest", "forest", "Where?", "OG0021_SC03_I", "YES", 2.0, "100% if placed in Where", "Correct story place/background.", "setting"],
    ["loves_colors", "loves changing colors", "At first...", "OG0021_SC01_I", "YES", 1.5, "100% if placed in At first", "Correct verb phrase from the opening state.", "setting"],
    ["butterfly", "butterfly", "Who?", "OG0021_SC03_I", "NO", 2.5, "35% slot credit if placed in Who", "Confuses first encountered character with Milo.", "setting"],
]
for r, row in enumerate(setting_rows, 5):
    fill = C_OK if row[4] == "YES" else C_LOW
    for c, v in enumerate(row, 1):
        bg = fill if c in (1, 5, 6) else None
        style_cell(ws.cell(r, c, v), bg, row[4] == "YES" and c in (1, 5, 6), "FFFFFF" if bg == C_OK and c in (1, 5, 6) else "000000")
    ws.row_dimensions[r].height = 30
ws.merge_cells("A13:I13")
style_cell(ws["A13"], C_GREEN, True).value = "Scoring: full slot weight for exact target; 35% of slot weight for same-category distractor; 0 for wrong category."

# Q03
make_option_sheet(
    "Q03 - Listening Scene Match | Story Grammar: Initiating Event",
    "Q03_INIT_EVENT",
    "Listen. Which scene starts the problem?",
    [
        ["A", "OG0021_SC02_I", "YES", 100, "Initiating Event", "Correct", "Milo wakes up gray; the story problem begins.", "-", "initiating_event"],
        ["B", "OG0021_SC03_I", "NO", 20, "Attempt", "Later action", "Student chooses the search after the problem, not the problem itself.", "Event/action sequence confusion", "attempt"],
        ["C", "OG0021_SC06_I", "NO", 30, "Reaction", "Result scene", "Student focuses on sadness after the problem.", "Initiating event vs reaction confusion", "reaction"],
        ["D", "OG0021_SC09_I", "NO", 0, "Resolution", "Opposite phase", "Student chooses a late recovery scene.", "Story direction weakness", "consequence"],
    ],
    "initiating_event",
    "Audio/OG0021_SC02_ST01_N_A.mp3",
)

# Q04
ws = wb.create_sheet("Q04_ATTEMPT")
set_widths(ws, [6, 18, 18, 14, 14, 16, 18, 24, 24])
merge_title(ws, "A1:I1", "Q04 - Scene-Anchored Unscramble | Story Grammar: Attempt")
ws.merge_cells("A2:I2")
style_cell(ws["A2"], C_YELLOW, True).value = "Question: Put the story words in order."
ws.merge_cells("A3:I3")
style_cell(ws["A3"], C_YELLOW).value = 'Correct sentence from SC03_ST01_N: "Milo walks into the forest."'
ws.merge_cells("A4:I4")
style_cell(ws["A4"], C_ACCENT).value = "Scrambled words: walks | forest. | Milo | the | into"
for i, h in enumerate(["Word", "Correct Pos", "Grammar Role", "Word Weight", "Weight Rationale", "Max Points", "Error Tag"], 1):
    header(ws, 6, i, h)
words = [
    ["Milo", 1, "Subject", 1.5, "Identifies actor of attempt.", "=D7/SUM($D$7:$D$11)*100", "actor_order"],
    ["walks", 2, "Action verb", 2.5, "Core attempt action; highest weight.", "=D8/SUM($D$7:$D$11)*100", "attempt_action"],
    ["into", 3, "Direction word", 1.5, "Connects action to place.", "=D9/SUM($D$7:$D$11)*100", "direction_word"],
    ["the", 4, "Article", 1.0, "Small syntax support.", "=D10/SUM($D$7:$D$11)*100", "article"],
    ["forest.", 5, "Place noun", 2.5, "Shows where the attempt begins.", "=D11/SUM($D$7:$D$11)*100", "attempt_place"],
]
for r, row in enumerate(words, 7):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), C_GREEN if c == 6 else (C_YELLOW if c in (1, 3) else None), False, "000000", 9, "center" if c in (2, 4, 6) else "left")
    ws.row_dimensions[r].height = 30
for c, v in [(1, "TOTAL"), (4, "=SUM(D7:D11)"), (6, "=SUM(F7:F11)")]:
    style_cell(ws.cell(12, c, v), C_ACCENT, True, "000000", 9, "center")
ws.merge_cells("A15:I15")
header(ws, 15, 1, "SECTION B - Word Position Score Matrix")
for i, h in enumerate(["Word \\ Submitted Pos", "Pos 1", "Pos 2", "Pos 3", "Pos 4", "Pos 5"], 1):
    header(ws, 16, i, h, C_ACCENT if i == 1 else C_SUB)
for idx, word in enumerate(words):
    source_row = 7 + idx
    out_row = 17 + idx
    correct_pos = int(word[1])
    style_cell(ws.cell(out_row, 1, word[0]), C_ACCENT, True, "000000", 9, "center")
    for pos in range(1, 6):
        fill = C_OK if pos == correct_pos else C_BAD
        font_color = "FFFFFF" if pos == correct_pos else "000000"
        formula = f"=IF({pos}=$B${source_row},ROUND($D${source_row}/SUM($D$7:$D$11)*100,1),0)"
        style_cell(ws.cell(out_row, pos + 1, formula), fill, pos == correct_pos, font_color, 9, "center")

ws.merge_cells("A24:I24")
header(ws, 24, 1, "SECTION C - Partial Score Examples")
for i, h in enumerate(["Student Answer", "Score", "Interpretation", "Weakness Signal"], 1):
    header(ws, 25, i, h, C_ACCENT)
examples = [
    ["Milo walks into the forest.", 100, "Exact story sentence restored.", "-"],
    ["Milo walks the into forest.", 83, "Core attempt understood; small syntax issue.", "function word placement"],
    ["forest. into the walks Milo", 17, "Place is known but action sequence is broken.", "attempt sequence weakness"],
]
for r, row in enumerate(examples, 26):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), option_fill(v) if c == 2 else None)
ws.merge_cells("A31:I31")
style_cell(ws["A31"], C_GREEN, True).value = 'LRS xAPI: verb="answered" | object="quiz_OG0021_q04_attempt" | result.sg_element="attempt" | result.word_order_submitted=[] | result.score_raw=<n> | result.word_scores={}'

# Q05-Q06
make_option_sheet(
    "Q05 - Feeling Match | Story Grammar: Reaction",
    "Q05_REACTION",
    "How does Milo feel here?",
    [
        ["A", "Happy", "NO", 0, "Reaction", "Opposite emotion", "Confuses resolution emotion with sadness.", "Temporal emotion confusion", "reaction"],
        ["B", "Sad", "YES", 100, "Reaction", "Correct", "SC06 shows crying and sadness.", "-", "reaction"],
        ["C", "Angry", "NO", 40, "Reaction", "Adjacent negative emotion", "Recognizes negative feeling but labels it too strongly.", "Emotion differentiation", "reaction"],
        ["D", "Surprised", "NO", 20, "Reaction", "Early-event emotion", "Confuses surprise with later sadness.", "Scene emotion mapping", "reaction"],
    ],
    "reaction",
    "OG0021_SC06_I.png",
)

make_option_sheet(
    "Q06 - Internal Response MCQ | Story Grammar: Internal Response",
    "Q06_INTERNAL",
    "What is Milo thinking?",
    [
        ["A", "Everyone has their own color.", "YES", 100, "Internal Response", "Correct", "Uses the actual SC06 thought/sentence to infer Milo's inner state.", "-", "internal_response"],
        ["B", "I want to fly with the butterfly.", "NO", 20, "Attempt detail", "Earlier action", "Confuses earlier butterfly scene with Milo's thought.", "Thought/action confusion", "attempt"],
        ["C", "The pond is very blue.", "NO", 45, "Surface observation", "Visual detail", "Sees visual detail but not internal state.", "Surface-level inference", "internal_response"],
        ["D", "I do not need my color.", "NO", 0, "Opposite motive", "Contradiction", "Contradicts Milo's motivation to find his color.", "Motive inversion", "internal_response"],
    ],
    "internal_response",
    "OG0021_SC06_I.png",
)

# SG_SCORING
ws = wb.create_sheet("SG_SCORING")
set_widths(ws, [18, 14, 18, 28, 34, 30, 28])
merge_title(ws, "A1:G1", "Story Grammar Scoring Model - v3")
for i, h in enumerate(["Axis", "Q_ID", "Score Source", "Calculation", "Interpretation", "Parent Report Use", "Weekly Rollup"], 1):
    header(ws, 3, i, h)
axes = [
    ["Setting", "Q02", "setting_score", "weighted slot match", "Builds who/where/at-first information from the first scene.", "Radar axis", "weighted average by story level"],
    ["Initiating Event", "Q03", "initiating_event_score", "option score", "Understands what started the problem.", "Radar axis", "weighted average by story level"],
    ["Attempt", "Q04", "attempt_score", "weighted word sequence", "Understands what the character does to solve the problem.", "Radar axis", "weighted average by story level"],
    ["Reaction", "Q05", "reaction_score", "option score", "Understands visible feeling/response.", "Radar axis", "weighted average by story level"],
    ["Internal Response", "Q06", "internal_response_score", "option score", "Infers thoughts, motive, inner state.", "Radar axis", "weighted average by story level"],
    ["Consequence", "Q01", "consequence_score", "weighted scene position", "Understands result and event development.", "Radar axis", "weighted average by story level"],
]
for r, row in enumerate(axes, 4):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), C_YELLOW if c == 1 else None)
ws.merge_cells("A13:G13")
style_cell(ws["A13"], C_GREEN, True).value = "Overall Reading Score = average(6 Story Grammar axes)"

# LRS_MAPPING
ws = wb.create_sheet("LRS_MAPPING")
set_widths(ws, [8, 20, 32, 24, 16, 16, 32, 34])
merge_title(ws, "A1:H1", "LRS xAPI Mapping - OG0021 v3")
for i, h in enumerate(["Q_ID", "xAPI Verb", "xAPI Object", "result.sg_element", "score_raw", "correct", "Extra Fields", "Risk Signal"], 1):
    header(ws, 2, i, h)
for r, row in enumerate(
    [
        ["Q01", "answered", "quiz_OG0021_v3_Q01_consequence", "consequence", "0-100", "partial", "scene_order, component_scores, hint_used", "Consequence gap if <70"],
        ["Q02", "answered", "quiz_OG0021_v3_Q02_setting", "setting", "0-100", "partial", "slot_values, component_scores, hint_used", "Setting gap if <70"],
        ["Q03", "answered", "quiz_OG0021_v3_Q03_initiating_event", "initiating_event", "0-100", "true/false", "option_selected, audio_src, hint_used", "Initiating event gap if <70"],
        ["Q04", "answered", "quiz_OG0021_v3_Q04_attempt", "attempt", "0-100", "partial", "word_order, word_scores, hint_used", "Attempt/action gap if <70"],
        ["Q05", "answered", "quiz_OG0021_v3_Q05_reaction", "reaction", "0-100", "true/false", "option_selected, hint_used", "Reaction/emotion gap if <70"],
        ["Q06", "answered", "quiz_OG0021_v3_Q06_internal_response", "internal_response", "0-100", "true/false", "option_selected, hint_used", "Inference gap if <70"],
    ],
    3,
):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), C_YELLOW if c == 1 else None)

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

wb.save(OUT)
print(f"Saved {OUT}")
