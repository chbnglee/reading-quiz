from __future__ import annotations

import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
RESOURCES = ROOT / "Resources_0708"
OUT_ROOT = ROOT / "v3"
BOOKEY_SRC = OUT_ROOT / "Assets" / "BKTK_Characters_Bookey.png"


SG_LABELS = {
    "consequence": ("Consequence", "결과 이해"),
    "setting": ("Setting", "배경 이해"),
    "initiating_event": ("Initiating Event", "사건 시작"),
    "attempt": ("Attempt", "해결 행동"),
    "reaction": ("Reaction", "감정 반응"),
    "internal_response": ("Internal Response", "내면 추론"),
}


@dataclass
class StoryDef:
    code: str
    title: str
    resource_dir: str
    level: str
    setting: dict[str, Any]
    sequence: list[str]
    sequence_titles: dict[str, str]
    event_scene: str
    event_sentence: str
    event_options: list[dict[str, Any]]
    attempt_scene: str
    attempt_sentence: str
    reaction_scene: str
    reaction_options: list[dict[str, Any]]
    internal_scene: str
    internal_options: list[dict[str, Any]]
    hints: list[str]
    setting_scene: str = "SC01"


def clean_text(text: str) -> str:
    text = text.replace("##", "")
    text = text.replace("“", '"').replace("”", '"').replace("’", "'")
    return text.strip()


def parse_story(path: Path) -> tuple[str, dict[str, list[dict[str, str]]], dict[str, str]]:
    text = path.read_text(encoding="utf-8")
    scenes: dict[str, list[dict[str, str]]] = {}
    emotions: dict[str, str] = {}
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        sent = re.match(r"^(SC\d{2})_(ST\d{2})_N\s*=\s*(.+)$", line)
        if sent:
            scene, st, value = sent.groups()
            scenes.setdefault(scene, []).append({
                "sentenceId": f"{scene}_{st}_N",
                "text": clean_text(value),
            })
            continue
        emo = re.match(r"^(SC\d{2})_Emotion\s*=\s*(.+)$", line)
        if emo:
            emotions[emo.group(1)] = emo.group(2).strip()
    return text, scenes, emotions


def word_tokens(sentence: str) -> list[str]:
    return re.findall(r"[A-Za-z']+[,\.!?]?", clean_text(sentence))


def ensure_dirs(base: Path) -> None:
    for name in ["Image", "Audio", "Cover", "Assets"]:
        (base / name).mkdir(parents=True, exist_ok=True)


def copy_first(srcs: list[Path], dst: Path) -> str:
    if not srcs:
        return ""
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(srcs[0], dst)
    return dst.name


def copy_named_asset(srcs: list[Path], dst_dir: Path, stem: str) -> str:
    if not srcs:
        return ""
    src = srcs[0]
    dst = dst_dir / f"{stem}{src.suffix.lower()}"
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return dst.name


def copy_assets(story: StoryDef, src_dir: Path, out_dir: Path) -> dict[str, Any]:
    ensure_dirs(out_dir)
    copied_images: dict[str, str] = {}
    image_src_dir = src_dir / "1080p"
    for img in sorted(image_src_dir.glob(f"{story.code}_SC*_I*.*")):
        if img.suffix.lower() not in [".webp", ".png", ".jpg", ".jpeg"]:
            continue
        scene_match = re.search(r"_(SC\d{2})_I", img.name)
        if not scene_match:
            continue
        scene = scene_match.group(1)
        ext = img.suffix.lower()
        dst_name = f"{story.code}_{scene}_I{ext}"
        shutil.copy2(img, out_dir / "Image" / dst_name)
        copied_images[scene] = dst_name

    audio_src_dir = src_dir / f"{story.code}_Audio_N_A"
    copied_audio: dict[str, str] = {}
    if audio_src_dir.exists():
        for audio in sorted(audio_src_dir.glob("*.mp3")):
            shutil.copy2(audio, out_dir / "Audio" / audio.name)
            stem = audio.stem.replace(f"{story.code}_", "").replace("_A", "")
            copied_audio[stem] = audio.name

    cover_name = copy_named_asset(
        [p for p in src_dir.glob(f"{story.code}_Cover_L_I*.*") if p.suffix.lower() in [".webp", ".png", ".jpg", ".jpeg"]],
        out_dir / "Cover",
        f"{story.code}_Cover_L_I",
    )
    bg_name = copy_named_asset(
        [p for p in src_dir.glob(f"{story.code}_Talking_BG_I*.*") if p.suffix.lower() in [".webp", ".png", ".jpg", ".jpeg"]],
        out_dir / "Image",
        f"{story.code}_Talking_BG_I",
    )
    if BOOKEY_SRC.exists():
        shutil.copy2(BOOKEY_SRC, out_dir / "Assets" / BOOKEY_SRC.name)

    return {
        "images": copied_images,
        "audio": copied_audio,
        "cover": cover_name,
        "background": bg_name,
        "bookey": BOOKEY_SRC.name if BOOKEY_SRC.exists() else "",
    }


def scene_image(assets: dict[str, Any], scene: str) -> str:
    return assets["images"].get(scene, f"{assets.get('code', '')}_{scene}_I.webp")


def polish_ko(text: str) -> str:
    replacements = {
        "혼동함": "혼동합니다.",
        "선택함": "선택합니다.",
        "적용함": "적용합니다.",
        "오해함": "오해합니다.",
        "우선함": "우선합니다.",
        "놓침": "놓칩니다.",
        "머무름": "머무릅니다.",
        "못함": "못합니다.",
        "투영함": "투영합니다.",
    }
    for old, new in replacements.items():
        if text.endswith(old):
            return text[: -len(old)] + new
    return text


def normalize_option(opt: dict[str, Any]) -> dict[str, Any]:
    out = {
        "key": opt["key"],
        "text": opt["text"],
        "score": opt["score"],
        "isCorrect": bool(opt.get("isCorrect")),
        "diagnostic": polish_ko(opt.get("diagnostic", "")),
    }
    if "scene" in opt:
        out["scene"] = opt["scene"]
    return out


def make_quiz(story: StoryDef, scenes: dict[str, list[dict[str, str]]], emotions: dict[str, str], assets: dict[str, Any]) -> dict[str, Any]:
    code = story.code
    assets["code"] = code
    character_names = {
        "OG0021": {"reaction": "Milo", "internal": "Milo"},
        "OG0036": {"reaction": "Judy", "internal": "Judy"},
        "OG0049": {"reaction": "the anglerfish", "internal": "Toby"},
        "CS0003": {"reaction": "Hans", "internal": "Hans"},
        "CS0006": {"reaction": "the Cat", "internal": "the Cat"},
        "OG0005": {"reaction": "Didi", "internal": "Didi"},
    }.get(code, {"reaction": "the character", "internal": "the character"})
    sequence_weights = {
        sc: (2.5 if idx in [0, len(story.sequence) - 1] else 1.5)
        for idx, sc in enumerate(story.sequence)
    }
    setting_slots = [
        {"key": "who", "label": "Who?", "correct": story.setting["who"]["correct"], "weight": 2.5},
        {"key": "where", "label": "Where?", "correct": story.setting["where"]["correct"], "weight": 2.0},
        {"key": "at_first", "label": "At first...", "correct": story.setting["at_first"]["correct"], "weight": 1.5},
    ]
    setting_items = []
    for slot_key, slot_meta in story.setting.items():
        for item in slot_meta["items"]:
            setting_items.append({
                "key": item["key"],
                "text": item["text"],
                "slot": slot_key,
                "diagnostic": polish_ko(item.get("diagnostic", "")),
            })

    attempt_words = word_tokens(story.attempt_sentence)
    word_weights = {
        word: (2.5 if idx in [0, 1, len(attempt_words) - 1] else 1.0)
        for idx, word in enumerate(attempt_words)
    }
    event_audio_stem = story.event_sentence.replace("_", "_")
    event_audio = assets["audio"].get(event_audio_stem, f"{code}_{story.event_sentence}_A.mp3")

    questions = [
        {
            "qId": f"{code}_V3_Q01",
            "number": 1,
            "type": "story_sequence_drag",
            "storyGrammar": "consequence",
            "instruction": "Put the story scenes in order.",
            "hint": story.hints[0],
            "resources": {"images": [{"id": sc, "path": scene_image(assets, sc), "kind": "image", "sceneId": sc, "caption": story.sequence_titles.get(sc, sc)} for sc in story.sequence]},
            "interaction": {"promptMode": "drag_sequence", "items": story.sequence[:], "correct": story.sequence[:]},
            "scoring": {
                "type": "weighted_position",
                "maxScore": 100,
                "formula": "score = round(sum(weight_i * max(0, 1 - abs(placed_pos_i - correct_pos_i) * 0.5)) / sum(weights) * 100)",
                "components": [{"key": sc, "weight": wt, "rule": "position_distance", "correctValue": idx + 1, "rationale": story.sequence_titles.get(sc, sc)} for idx, (sc, wt) in enumerate(sequence_weights.items())],
            },
            "diagnostics": [{"code": "consequence_sequence_gap", "threshold": 70, "messageKo": "사건의 흐름과 결과를 순서대로 다시 확인하는 연습이 필요합니다."}],
            "lrs": {"verb": "answered", "objectId": f"quiz_{code}_v3_Q01_consequence", "resultFields": ["score_raw", "scene_order", "hint_used"]},
        },
        {
            "qId": f"{code}_V3_Q02",
            "number": 2,
            "type": "setting_slot_drag",
            "storyGrammar": "setting",
            "instruction": "Look at the picture. Fill in the boxes.",
            "hint": story.hints[1],
            "resources": {"images": [{"id": story.setting_scene, "path": scene_image(assets, story.setting_scene), "kind": "image", "sceneId": story.setting_scene}], "scene": story.setting_scene},
            "interaction": {"promptMode": "slot_drag", "slots": setting_slots, "items": setting_items, "correct": {s["key"]: s["correct"] for s in setting_slots}},
            "scoring": {
                "type": "weighted_slot_match",
                "maxScore": 100,
                "formula": "score = round(sum(slot_weight * (1 if exact card else .35 if same slot category else 0)) / sum(weights) * 100)",
                "components": [{"key": s["key"], "weight": s["weight"], "rule": "slot_match", "correctValue": s["correct"], "partialCredit": 0.35, "rationale": f"Identifies {s['label']} in the opening setting."} for s in setting_slots],
            },
            "diagnostics": [{"code": "setting_gap", "threshold": 70, "messageKo": "이야기의 처음 장면에서 인물, 장소, 처음 상황을 구분하는 연습이 필요합니다."}],
            "lrs": {"verb": "answered", "objectId": f"quiz_{code}_v3_Q02_setting", "resultFields": ["score_raw", "slot_values", "hint_used"]},
        },
        {
            "qId": f"{code}_V3_Q03",
            "number": 3,
            "type": "listen_scene_mcq",
            "storyGrammar": "initiating_event",
            "instruction": "Listen. Which scene starts the problem?",
            "hint": story.hints[2],
            "resources": {
                "images": [{"id": opt["scene"], "path": scene_image(assets, opt["scene"]), "kind": "image", "sceneId": opt["scene"]} for opt in story.event_options],
                "audio": {"id": f"{story.event_sentence}_A", "path": event_audio, "kind": "audio", "sceneId": story.event_scene, "sentenceId": story.event_sentence},
            },
            "interaction": {"promptMode": "image_mcq", "options": [normalize_option(o) for o in story.event_options], "correct": next(o["key"] for o in story.event_options if o.get("isCorrect"))},
            "scoring": {
                "type": "fixed_option_score",
                "maxScore": 100,
                "formula": "score = selected_option.score",
                "components": [{"key": o["key"], "weight": o["score"], "rule": "option_score", "correctValue": bool(o.get("isCorrect")), "rationale": polish_ko(o.get("diagnostic", "Correct initiating event scene."))} for o in story.event_options],
            },
            "diagnostics": [{"code": "initiating_event_gap", "threshold": 70, "messageKo": "문제가 시작되는 장면과 다른 장면을 구분하는 연습이 필요합니다."}],
            "lrs": {"verb": "answered", "objectId": f"quiz_{code}_v3_Q03_initiating_event", "resultFields": ["score_raw", "option_selected", "audio_src", "hint_used"]},
        },
        {
            "qId": f"{code}_V3_Q04",
            "number": 4,
            "type": "scene_word_unscramble",
            "storyGrammar": "attempt",
            "instruction": "Put the story words in order.",
            "hint": story.hints[3],
            "resources": {"images": [{"id": story.attempt_scene, "path": scene_image(assets, story.attempt_scene), "kind": "image", "sceneId": story.attempt_scene}], "scene": story.attempt_scene},
            "interaction": {"promptMode": "word_unscramble", "items": list(reversed(attempt_words)), "correct": attempt_words},
            "scoring": {
                "type": "weighted_word_position",
                "maxScore": 100,
                "formula": "score = round(sum(weight[word] if submitted_pos == correct_pos) / sum(weights) * 100)",
                "components": [{"key": word, "weight": weight, "rule": "exact_position", "correctValue": idx + 1, "rationale": "Actor/action/result word order."} for idx, (word, weight) in enumerate(word_weights.items())],
            },
            "diagnostics": [{"code": "attempt_sentence_gap", "threshold": 70, "messageKo": "인물이 문제를 해결하려고 한 행동 문장을 순서대로 구성하는 연습이 필요합니다."}],
            "lrs": {"verb": "answered", "objectId": f"quiz_{code}_v3_Q04_attempt", "resultFields": ["score_raw", "word_order", "hint_used"]},
        },
        {
            "qId": f"{code}_V3_Q05",
            "number": 5,
            "type": "emotion_mcq",
            "storyGrammar": "reaction",
            "instruction": f"How does {character_names['reaction']} feel here?",
            "hint": story.hints[4],
            "resources": {"images": [{"id": story.reaction_scene, "path": scene_image(assets, story.reaction_scene), "kind": "image", "sceneId": story.reaction_scene}], "scene": story.reaction_scene, "emotionSource": emotions.get(story.reaction_scene, "")},
            "interaction": {"promptMode": "text_mcq", "options": [normalize_option(o) for o in story.reaction_options], "correct": next(o["key"] for o in story.reaction_options if o.get("isCorrect"))},
            "scoring": {
                "type": "fixed_option_score",
                "maxScore": 100,
                "formula": "score = selected_option.score",
                "components": [{"key": o["key"], "weight": o["score"], "rule": "option_score", "correctValue": bool(o.get("isCorrect")), "rationale": polish_ko(o.get("diagnostic", "Correct reaction/emotion."))} for o in story.reaction_options],
            },
            "diagnostics": [{"code": "reaction_emotion_gap", "threshold": 70, "messageKo": "장면의 사건과 인물의 감정 반응을 연결하는 연습이 필요합니다."}],
            "lrs": {"verb": "answered", "objectId": f"quiz_{code}_v3_Q05_reaction", "resultFields": ["score_raw", "option_selected", "hint_used"]},
        },
        {
            "qId": f"{code}_V3_Q06",
            "number": 6,
            "type": "internal_response_mcq",
            "storyGrammar": "internal_response",
            "instruction": f"What is {character_names['internal']} thinking?",
            "hint": story.hints[5],
            "resources": {"images": [{"id": story.internal_scene, "path": scene_image(assets, story.internal_scene), "kind": "image", "sceneId": story.internal_scene}], "scene": story.internal_scene},
            "interaction": {"promptMode": "text_mcq", "options": [normalize_option(o) for o in story.internal_options], "correct": next(o["key"] for o in story.internal_options if o.get("isCorrect"))},
            "scoring": {
                "type": "fixed_option_score",
                "maxScore": 100,
                "formula": "score = selected_option.score",
                "components": [{"key": o["key"], "weight": o["score"], "rule": "option_score", "correctValue": bool(o.get("isCorrect")), "rationale": polish_ko(o.get("diagnostic", "Correct internal response."))} for o in story.internal_options],
            },
            "diagnostics": [{"code": "internal_response_gap", "threshold": 70, "messageKo": "인물의 생각, 의도, 깨달음을 장면 근거로 추론하는 연습이 필요합니다."}],
            "lrs": {"verb": "answered", "objectId": f"quiz_{code}_v3_Q06_internal_response", "resultFields": ["score_raw", "option_selected", "hint_used"]},
        },
    ]
    return {
        "schemaVersion": "quiz-v3.0",
        "story": {
            "storyId": code,
            "title": story.title,
            "level": story.level,
            "text": "\n".join(f"{scene}_{s['sentenceId'].split('_', 1)[1]} = {s['text']}" for scene, sent_list in scenes.items() for s in sent_list),
            "scenes": [{"sceneId": scene, "sentences": sent_list} for scene, sent_list in scenes.items()],
        },
        "assets": {
            "imageBasePath": "Image/",
            "audioBasePath": "Audio/",
            "coverBasePath": "Cover/",
            "backgroundImage": f"Image/{assets['background']}" if assets.get("background") else "",
            "coverImage": f"Cover/{assets['cover']}" if assets.get("cover") else "",
            "hintCharacter": f"Assets/{assets['bookey']}" if assets.get("bookey") else "",
        },
        "storyGrammarAxes": [{"key": key, "labelEn": en, "labelKo": ko} for key, (en, ko) in SG_LABELS.items()],
        "questions": questions,
        "reporting": {
            "overallFormula": "overall = average(consequence, setting, initiating_event, attempt, reaction, internal_response)",
            "masteryBands": [
                {"key": "stable", "min": 85, "max": 100, "labelKo": "안정"},
                {"key": "developing", "min": 70, "max": 84, "labelKo": "발달 중"},
                {"key": "shaky", "min": 50, "max": 69, "labelKo": "흔들림"},
                {"key": "focus", "min": 0, "max": 49, "labelKo": "집중 보완"},
            ],
        },
        "generation": {"provider": "codex", "model": "manual-v3-logic", "promptVersion": "story_grammar_v3", "createdAt": "2026-07-08", "notes": "Generated directly from Resources_0708."},
    }


HEADER = PatternFill("solid", fgColor="1F4E79")
SUBHEADER = PatternFill("solid", fgColor="D9D9D9")
LIGHT_BLUE = PatternFill("solid", fgColor="BDD7EE")
LIGHT_YELLOW = PatternFill("solid", fgColor="FFF2CC")
LIGHT_GREEN = PatternFill("solid", fgColor="E2EFDA")
SECTION_RESOURCES = PatternFill("solid", fgColor="4C1D95")
SECTION_INTERACTION = PatternFill("solid", fgColor="4C1D95")
SECTION_SCORING = PatternFill("solid", fgColor="4C1D95")
SECTION_DIAGNOSTICS = PatternFill("solid", fgColor="4C1D95")
SECTION_RUBRIC = PatternFill("solid", fgColor="4C1D95")
GREEN = PatternFill("solid", fgColor="70AD47")
YELLOW = PatternFill("solid", fgColor="FFD966")
LOW = PatternFill("solid", fgColor="FFEECC")
RED = PatternFill("solid", fgColor="FFCCCC")
WHITE = "FFFFFF"
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=9)
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    accent_colors = {"004C1D95", "00D9D9D9"}
    for row in ws.iter_rows():
        first = row[0].value
        if isinstance(first, str) and (first.startswith("SECTION") or first in {"Kind", "Key", "Slot", "Items", "Code", "Scene", "Card", "Example"}):
            for cell in row:
                rgb = getattr(cell.fill.fgColor, "rgb", None)
                if rgb in accent_colors or first.startswith("SECTION"):
                    cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE if first.startswith("SECTION") else "000000")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(42, max(12, ws.column_dimensions[get_column_letter(col)].width or 12))


def append_title(ws, title: str, cols: int) -> None:
    ws.append([title] + [""] * (cols - 1))
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=cols)


def append_header(ws, values: list[Any]) -> None:
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.fill = SUBHEADER
        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")


def append_section(ws, values: list[Any], fill: PatternFill) -> None:
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.fill = fill
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)


def score_fill(score: float) -> PatternFill:
    if score >= 100:
        return GREEN
    if score >= 30:
        return YELLOW
    if score > 0:
        return LOW
    return RED


def build_reading_xlsx(quiz: dict[str, Any], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "QUIZ_LIST"
    append_title(ws, f"{quiz['story']['storyId']} Reading Quiz v3 ({quiz['story']['level']})", 9)
    append_header(ws, ["Q_ID", "No", "Story Grammar", "Question Type", "Instruction", "Hint", "Max Score", "Formula", "LRS Object"])
    for q in quiz["questions"]:
        ws.append([q["qId"], q["number"], q["storyGrammar"], q["type"], q["instruction"], q["hint"], q["scoring"]["maxScore"], q["scoring"]["formula"], q["lrs"]["objectId"]])
    style_sheet(ws)
    for q in quiz["questions"]:
        sheet = wb.create_sheet(f"Q{q['number']:02d}_{q['storyGrammar']}"[:31].upper())
        append_title(sheet, f"Q{q['number']:02d} - {q['storyGrammar']} / {q['type']}", 8)
        sheet.append(["Q_ID", q["qId"], "Instruction", q["instruction"], "Hint", q["hint"], "Story Grammar", q["storyGrammar"]])
        sheet.append(["Scoring Type", q["scoring"]["type"], "Formula", q["scoring"]["formula"], "Max Score", q["scoring"]["maxScore"], "", ""])
        append_section(sheet, ["SECTION A", "Resources", "", "", "", "", "", ""], SECTION_RESOURCES)
        append_header(sheet, ["Kind", "ID", "Path", "Scene ID", "Sentence ID", "Caption/Text", "", ""])
        for img in q["resources"].get("images", []):
            sheet.append(["image", img.get("id", ""), img.get("path", ""), img.get("sceneId", ""), img.get("sentenceId", ""), img.get("caption", ""), "", ""])
        if q["resources"].get("audio"):
            a = q["resources"]["audio"]
            sheet.append(["audio", a.get("id", ""), a.get("path", ""), a.get("sceneId", ""), a.get("sentenceId", ""), "", "", ""])
        append_section(sheet, ["SECTION B", "Interaction", "", "", "", "", "", ""], SECTION_INTERACTION)
        if q["interaction"].get("options"):
            append_header(sheet, ["Key", "Text", "Score", "Correct?", "Diagnostic", "", "", ""])
            for opt in q["interaction"]["options"]:
                sheet.append([opt["key"], opt["text"], opt["score"], "YES" if opt.get("isCorrect") else "NO", opt.get("diagnostic", ""), "", "", ""])
                for cell in sheet[sheet.max_row][:4]:
                    cell.fill = score_fill(float(opt["score"]))
                    if opt["score"] >= 100:
                        cell.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
        elif q["interaction"].get("slots"):
            append_header(sheet, ["Slot", "Label", "Correct Card", "Weight", "Partial Credit", "", "", ""])
            for slot in q["interaction"]["slots"]:
                sheet.append([slot["key"], slot["label"], slot["correct"], slot["weight"], "35% same category", "", "", ""])
        else:
            append_header(sheet, ["Items", json.dumps(q["interaction"].get("items", []), ensure_ascii=False), "Correct", json.dumps(q["interaction"].get("correct", []), ensure_ascii=False), "", "", "", ""])
        append_section(sheet, ["SECTION C", "Scoring Components", "", "", "", "", "", ""], SECTION_SCORING)
        append_header(sheet, ["Key", "Weight/Score", "Rule", "Correct Value", "Partial Credit", "Rationale", "", ""])
        for comp in q["scoring"].get("components", []):
            sheet.append([comp.get("key", ""), comp.get("weight", ""), comp.get("rule", ""), comp.get("correctValue", ""), comp.get("partialCredit", ""), comp.get("rationale", ""), "", ""])
            if isinstance(comp.get("weight"), (int, float)):
                sheet.cell(sheet.max_row, 2).fill = score_fill(float(comp["weight"]) if q["scoring"]["type"] == "fixed_option_score" else 100)
        if q["type"] == "story_sequence_drag":
            append_section(sheet, ["SECTION E", "Position Rubric: Submitted Position Score", "", "", "", "", "", ""], SECTION_RUBRIC)
            pos_count = len(q["interaction"]["correct"])
            total_weight = sum(float(c.get("weight", 0)) for c in q["scoring"].get("components", [])) or 1
            append_header(sheet, ["Scene", "Correct Pos", "Weight", "Max Points"] + [f"Pos {idx}" for idx in range(1, pos_count + 1)])
            for comp in q["scoring"].get("components", []):
                correct_pos = int(comp.get("correctValue", 0))
                max_points = float(comp.get("weight", 0)) / total_weight * 100
                row = [comp.get("key", ""), correct_pos, comp.get("weight", ""), round(max_points, 1)]
                for submitted_pos in range(1, pos_count + 1):
                    distance = abs(submitted_pos - correct_pos)
                    row.append(round(max_points * max(0, 1 - distance * 0.5), 1))
                sheet.append(row)
                for idx in range(5, 5 + pos_count):
                    val = sheet.cell(sheet.max_row, idx).value or 0
                    if abs(float(val) - max_points) < 0.1:
                        sheet.cell(sheet.max_row, idx).fill = GREEN
                        sheet.cell(sheet.max_row, idx).font = Font(name="Calibri", size=9, bold=True, color=WHITE)
                    elif float(val) > 0:
                        sheet.cell(sheet.max_row, idx).fill = YELLOW
                    else:
                        sheet.cell(sheet.max_row, idx).fill = RED
        if q["type"] == "setting_slot_drag":
            append_section(sheet, ["SECTION E", "Slot/Card Credit Matrix", "", "", "", "", "", ""], SECTION_RUBRIC)
            slots = q["interaction"].get("slots", [])
            items = q["interaction"].get("items", [])
            total_weight = sum(float(s.get("weight", 0)) for s in slots) or 1
            append_header(sheet, ["Card", "Correct Category"] + [s.get("label", s.get("key", "")) for s in slots] + ["Rule"])
            for item in items:
                row = [item.get("text", ""), item.get("slot", "")]
                for slot in slots:
                    slot_points = float(slot.get("weight", 0)) / total_weight * 100
                    if item.get("key") == slot.get("correct"):
                        row.append(f"{round(slot_points, 1)} pts (full)")
                    elif item.get("slot") == slot.get("key"):
                        row.append(f"{round(slot_points * 0.35, 1)} pts (35%)")
                    else:
                        row.append("0 pts")
                row.append("Full credit for exact card; 35% for same category distractor; 0 for wrong category.")
                sheet.append(row)
        append_section(sheet, ["SECTION D", "Diagnostics", "", "", "", "", "", ""], SECTION_DIAGNOSTICS)
        append_header(sheet, ["Code", "Threshold", "Message", "", "", "", "", ""])
        for diag in q.get("diagnostics", []):
            sheet.append([diag["code"], diag["threshold"], diag["messageKo"], "", "", "", "", ""])
        style_sheet(sheet)
        widths = [20, 24, 28, 18, 20, 36, 16, 20]
        for idx, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(idx)].width = width
    sg = wb.create_sheet("SG_SCORING")
    append_title(sg, "Story Grammar Scoring", 5)
    append_header(sg, ["Axis", "Question", "Score Source", "Formula", "Report Use"])
    for q in quiz["questions"]:
        sg.append([q["storyGrammar"], q["qId"], "question_score", q["scoring"]["formula"], "radar/profile"])
    sg.append(["overall", "", "", quiz["reporting"]["overallFormula"], "parent report"])
    style_sheet(sg)
    lrs = wb.create_sheet("LRS_MAPPING")
    append_title(lrs, "LRS Mapping", 5)
    append_header(lrs, ["Q_ID", "Verb", "Object ID", "Story Grammar", "Result Fields"])
    for q in quiz["questions"]:
        lrs.append([q["qId"], q["lrs"]["verb"], q["lrs"]["objectId"], q["storyGrammar"], ", ".join(q["lrs"]["resultFields"])])
    style_sheet(lrs)
    wb.save(out_path)


def build_dev_xlsx(quiz: dict[str, Any], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "QUESTIONS"
    append_header(ws, ["q_id", "story_id", "story_level", "number", "story_grammar", "question_type", "instruction", "hint", "max_score", "formula"])
    for q in quiz["questions"]:
        ws.append([q["qId"], quiz["story"]["storyId"], quiz["story"]["level"], q["number"], q["storyGrammar"], q["type"], q["instruction"], q["hint"], q["scoring"]["maxScore"], q["scoring"]["formula"]])
    style_sheet(ws)
    res = wb.create_sheet("RESOURCES")
    append_header(res, ["q_id", "resource_kind", "resource_id", "path", "scene_id", "sentence_id"])
    for q in quiz["questions"]:
        for img in q["resources"].get("images", []):
            res.append([q["qId"], "image", img.get("id", ""), img.get("path", ""), img.get("sceneId", ""), img.get("sentenceId", "")])
        if q["resources"].get("audio"):
            a = q["resources"]["audio"]
            res.append([q["qId"], "audio", a.get("id", ""), a.get("path", ""), a.get("sceneId", ""), a.get("sentenceId", "")])
    style_sheet(res)
    opt = wb.create_sheet("OPTIONS")
    append_header(opt, ["q_id", "option_key", "option_text", "score", "is_correct", "diagnostic"])
    for q in quiz["questions"]:
        if q["interaction"].get("options"):
            for o in q["interaction"]["options"]:
                opt.append([q["qId"], o["key"], o["text"], o["score"], bool(o.get("isCorrect")), o.get("diagnostic", "")])
        elif q["interaction"].get("items"):
            for idx, item in enumerate(q["interaction"]["items"], 1):
                text = item["text"] if isinstance(item, dict) else item
                opt.append([q["qId"], str(idx), text, "", "", ""])
    style_sheet(opt)
    rules = wb.create_sheet("SCORING_RULES")
    append_header(rules, ["q_id", "component_key", "weight", "rule", "correct_value", "partial_credit", "rationale"])
    for q in quiz["questions"]:
        for c in q["scoring"].get("components", []):
            rules.append([q["qId"], c.get("key", ""), c.get("weight", ""), c.get("rule", ""), c.get("correctValue", ""), c.get("partialCredit", ""), c.get("rationale", "")])
    style_sheet(rules)
    lrs = wb.create_sheet("LRS_MAPPING")
    append_header(lrs, ["q_id", "verb", "object_id", "result_fields"])
    for q in quiz["questions"]:
        lrs.append([q["qId"], q["lrs"]["verb"], q["lrs"]["objectId"], "|".join(q["lrs"]["resultFields"])])
    style_sheet(lrs)
    wb.save(out_path)


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title} Reading Quiz</title>
<link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=ABeeZee&display=swap" rel="stylesheet">
<style>
:root{{--purple:#7C3AED;--ink:#1F2937;--soft:#F5F3FF;--line:#E5E7EB}}
*{{box-sizing:border-box}}
body{{margin:0;font-family:'ABeeZee',sans-serif;color:var(--ink);background:#F5F3FF}}
.screen{{display:none;min-height:100vh;padding:24px;align-items:center;justify-content:center;background-size:cover;background-position:center;background-attachment:fixed}}
.screen.active{{display:flex}}
.cover{{background:#fff}}
.cover-card{{width:min(920px,94vw);text-align:center}}
.cover-img{{width:min(760px,92vw);max-height:64vh;object-fit:contain;border-radius:28px;box-shadow:0 20px 60px rgba(31,41,55,.12)}}
h1{{font-family:'Nunito',sans-serif;font-size:42px;margin:22px 0 6px;color:#4C1D95}}
.start-btn,.btn{{border:none;border-radius:999px;padding:13px 28px;font-family:'Nunito',sans-serif;font-weight:900;cursor:pointer}}
.start-btn{{background:#7C3AED;color:white;font-size:19px;margin-top:20px}}
.top-nav{{position:fixed;top:12px;left:50%;transform:translateX(-50%);z-index:50;display:flex;gap:8px;background:rgba(255,255,255,.9);border:1px solid #EDE9FE;border-radius:999px;padding:8px 10px;box-shadow:0 10px 28px rgba(0,0,0,.08)}}
.dot{{width:32px;height:32px;border-radius:50%;border:2px solid #DDD6FE;background:white;color:#6B7280;font-family:'Nunito',sans-serif;font-weight:900}}
.dot.active{{background:#EDE9FE;color:#7C3AED;border-color:#7C3AED}}
.dot.done{{background:#7C3AED;color:white;border-color:#7C3AED}}
.q-card{{width:min(940px,96vw);background:rgba(255,255,255,.96);border-radius:28px;padding:24px;box-shadow:0 20px 60px rgba(31,41,55,.18)}}
.q-head{{display:flex;align-items:center;gap:12px;margin-bottom:18px}}
.badge{{width:46px;height:46px;border-radius:50%;display:flex;align-items:center;justify-content:center;background:#7C3AED;color:white;font-family:'Nunito',sans-serif;font-weight:900}}
.instruction{{font-size:20px;font-weight:800;line-height:1.35;flex:1}}
.tag{{border-radius:999px;padding:7px 13px;background:#EDE9FE;color:#7C3AED;font-family:'Nunito',sans-serif;font-weight:900;font-size:13px}}
.scene-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin:14px 0}}
.scene-grid.single-grid{{grid-template-columns:minmax(260px,480px);justify-content:center}}
.scene-grid.event-grid{{grid-template-columns:repeat(2,minmax(220px,1fr));max-width:760px;margin:14px auto}}
.scene-card{{border:2px solid #EDE9FE;border-radius:18px;overflow:hidden;background:white;cursor:pointer;transition:.15s;min-height:150px}}
.scene-card:disabled{{cursor:default}}
.scene-card.selected{{border-color:#7C3AED;box-shadow:0 0 0 4px #EDE9FE}}
.scene-card.correct{{border-color:#10B981;background:#ECFDF5}}
.scene-card.wrong{{border-color:#EF4444;background:#FEF2F2}}
.scene-card img{{width:100%;height:180px;object-fit:contain;display:block;background:#F9FAFB}}
.scene-grid.event-grid .scene-card img{{height:210px}}
.scene-card.image-only img{{height:160px}}
.scene-grid.single-grid .scene-card img{{height:240px;object-fit:cover}}
.scene-card .cap{{padding:8px;text-align:center;font-family:'Nunito',sans-serif;font-size:12px;color:#4B5563}}
.scene-card.image-only .cap{{display:none}}
.slots{{display:grid;grid-template-columns:repeat(auto-fit,minmax(126px,1fr));gap:10px;margin:14px 0}}
.slot{{min-height:94px;border:2px dashed #C4B5FD;border-radius:18px;background:#FAF7FF;display:flex;align-items:center;justify-content:center;text-align:center;padding:8px;color:#7C3AED;font-family:'Nunito',sans-serif;font-weight:900;cursor:pointer;overflow:hidden}}
.slot img{{width:100%;height:100%;object-fit:cover;border-radius:14px}}
.slot.filled{{border:2px solid #A78BFA;background:white;color:#4C1D95;box-shadow:0 0 0 4px #F5F3FF}}
.bank,.word-bank{{display:flex;flex-wrap:wrap;gap:10px;justify-content:center;margin:14px 0;padding:14px;background:#FFFBEB;border:2px solid #FDE68A;border-radius:20px}}
.seq-slots,.seq-bank{{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));max-width:840px;margin:14px auto}}
.seq-slots .slot,.seq-bank .scene-card{{grid-column:span 2}}
.seq-slots .slot:nth-child(4),.seq-bank .scene-card:nth-child(4){{grid-column:2 / span 2}}
.seq-slots .slot:nth-child(5),.seq-bank .scene-card:nth-child(5){{grid-column:4 / span 2}}
.setting-bank{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr))}}
.mini-card,.word{{border:2px solid #FCD34D;border-radius:999px;background:white;padding:11px 16px;font-family:'Nunito',sans-serif;font-weight:900;cursor:pointer;text-align:center}}
.word.correct{{border-color:#10B981;background:#ECFDF5;color:#065F46}}
.word.wrong{{border-color:#EF4444;background:#FEF2F2;color:#991B1B}}
.answer-row{{min-height:70px;border:2px dashed #C4B5FD;border-radius:18px;background:#FAF7FF;display:flex;align-items:center;justify-content:center;gap:8px;flex-wrap:wrap;padding:12px}}
.option-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px;margin:14px 0}}
.option{{border:2px solid #E5E7EB;border-radius:18px;background:white;padding:16px;text-align:left;cursor:pointer;font-size:16px;line-height:1.4;display:flex;align-items:center;gap:10px}}
.option.selected{{border-color:#7C3AED;background:#EDE9FE}}
.option.correct{{border-color:#10B981;background:#ECFDF5}}
.option.wrong{{border-color:#EF4444;background:#FEF2F2}}
.opt-letter{{font-family:'Nunito',sans-serif;font-weight:900;color:#7C3AED;font-size:19px;flex:0 0 auto}}
.audio-btn{{width:68px;height:68px;border-radius:50%;border:none;background:#7C3AED;color:white;font-size:26px;cursor:pointer;display:block;margin:6px auto 14px}}
.nav-row{{display:flex;justify-content:center;gap:10px;margin-top:18px}}
.btn-check{{background:#7C3AED;color:white}}.btn-next,.btn-retry{{background:#EDE9FE;color:#7C3AED;display:none}}.show{{display:inline-block}}.btn:disabled{{opacity:.45;cursor:not-allowed}}
.feedback{{display:none;margin-top:12px;text-align:center;font-family:'Nunito',sans-serif;font-weight:900;padding:10px;border-radius:14px}}.feedback.show{{display:block}}.feedback.ok{{background:#D1FAE5;color:#065F46}}.feedback.no{{background:#FEE2E2;color:#991B1B}}
.bookey{{position:fixed;left:18px;bottom:18px;z-index:80;display:none;align-items:flex-end;gap:10px}}.bookey.show{{display:flex}}
.bookey button{{width:58px;height:58px;border-radius:50%;border:2px solid #EDE9FE;background:white;padding:1px;overflow:hidden;cursor:pointer;box-shadow:0 10px 26px rgba(124,58,237,.16)}}.bookey img{{width:100%;height:100%;object-fit:contain}}
.bubble{{display:none;max-width:260px;background:white;border:2px solid #C4B5FD;border-radius:20px 20px 20px 6px;padding:12px 14px;font-size:14px;line-height:1.45;box-shadow:0 12px 28px rgba(31,41,55,.12)}}.bookey.open .bubble{{display:block}}
.download-box{{position:fixed;right:18px;bottom:18px;z-index:85}}
.download-main{{width:48px;height:48px;border:2px solid #EDE9FE;border-radius:50%;background:white;color:#7C3AED;font-size:22px;box-shadow:0 10px 26px rgba(124,58,237,.16);cursor:pointer}}
.download-menu{{display:none;position:absolute;right:0;bottom:48px;min-width:190px;background:white;border:2px solid #EDE9FE;border-radius:18px;padding:8px;box-shadow:0 14px 34px rgba(31,41,55,.16)}}
.download-box.open .download-menu{{display:block}}
.download-menu a{{display:block;padding:10px 12px;border-radius:12px;text-decoration:none;color:#4C1D95;font-family:'Nunito',sans-serif;font-weight:900;font-size:13px}}
.download-menu a:hover{{background:#F5F3FF}}
.result-card{{width:min(960px,96vw);background:white;border-radius:28px;padding:28px;box-shadow:0 20px 60px rgba(31,41,55,.12)}}
.ox-grid{{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin:22px 0}}.ox{{border:2px solid #E5E7EB;border-radius:16px;padding:12px;text-align:center;font-family:'Nunito',sans-serif;font-weight:900}}.ox.ok{{background:#ECFDF5;border-color:#34D399}}.ox.no{{background:#FEF2F2;border-color:#F87171}}
.parent-grid{{display:grid;grid-template-columns:360px 1fr;gap:18px;align-items:stretch}}canvas{{width:100%;height:300px;background:#F9FAFB;border:2px solid #EDE9FE;border-radius:20px}}
.score-list{{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}}.score-card{{border:2px solid #E5E7EB;border-radius:16px;padding:12px;background:#fff}}.score-card strong{{display:block;color:#4C1D95;margin-bottom:6px}}.score-card .score{{font-family:'Nunito',sans-serif;font-weight:900;font-size:22px}}
@media(max-width:760px){{.option-grid,.parent-grid,.score-list{{grid-template-columns:1fr}}.ox-grid{{grid-template-columns:repeat(3,1fr)}}}}
</style>
</head>
<body>
<div class="top-nav" id="topNav"></div>
<div id="cover" class="screen cover active"><div class="cover-card"><img class="cover-img" src="{cover}" alt="cover"><h1>Reading Quiz</h1><button class="start-btn" onclick="go(0)">Start</button></div></div>
<div id="screens"></div>
<div id="student" class="screen"><div class="result-card"><h1>Great job! 🌟</h1><p class="instruction">You finished the quiz.</p><div class="ox-grid" id="oxGrid"></div><button class="start-btn" onclick="showParent()">Parent Report</button></div></div>
<div id="parent" class="screen"><div class="result-card"><h1>학부모 리포트</h1><div class="parent-grid"><canvas id="radar" width="360" height="300"></canvas><div class="score-list" id="scoreList"></div></div><div class="nav-row"><button class="btn btn-next show" onclick="restart()">Restart</button></div></div></div>
<div class="bookey" id="bookey"><button onclick="toggleHint()"><img src="{bookey}" alt="Bookey"></button><div class="bubble" id="bubble"></div></div>
<div class="download-box" id="downloadBox"><button class="download-main" onclick="toggleDownload()" title="Download files">📥</button><div class="download-menu"><a href="{story_id}_ReadingQuiz.xlsx" download>Reading Quiz XLSX</a><a href="{story_id}_DevSpec.xlsx" download>DevSpec XLSX</a></div></div>
<script>
const QUIZ = {quiz_json};
const bg = QUIZ.assets.backgroundImage;
const scores = Array(QUIZ.questions.length).fill(null);
const answers = Array(QUIZ.questions.length).fill(null);
const answerDetails = Array(QUIZ.questions.length).fill(null);
let current = -1;
let selected = null;
let dragged = null;
const sgOrder = ["consequence","setting","initiating_event","attempt","reaction","internal_response"];
const sgNames = {{consequence:'Consequence',setting:'Setting',initiating_event:'Initiating Event',attempt:'Attempt',reaction:'Reaction',internal_response:'Internal Response'}};
function el(id){{return document.getElementById(id)}}
function img(path){{return QUIZ.assets.imageBasePath + path}}
function audio(path){{return QUIZ.assets.audioBasePath + path}}
function allScreens(){{return [el('cover'),...document.querySelectorAll('.qscreen'),el('student'),el('parent')]}}
function go(i){{allScreens().forEach(s=>s.classList.remove('active')); current=i; el('q'+i).classList.add('active'); el('bookey').classList.add('show'); updateNav(); updateHint();}}
function updateNav(){{const nav=el('topNav');nav.innerHTML='';QUIZ.questions.forEach((q,i)=>{{const b=document.createElement('button');b.className='dot '+(i===current?'active ':'')+(scores[i]!=null?'done':'');b.textContent=i+1;b.onclick=()=>go(i);nav.appendChild(b);}})}}
function updateHint(){{const q=QUIZ.questions[current];el('bubble').textContent=q?q.hint:'';el('bookey').classList.remove('open')}}
function toggleHint(){{el('bookey').classList.toggle('open')}}
function toggleDownload(){{el('downloadBox').classList.toggle('open')}}
function allowDrop(e){{e.preventDefault()}}
function dragStart(e,kind,i,value,node=null){{dragged={{kind,i,value}};if(e&&e.dataTransfer){{e.dataTransfer.setData('text/plain',value);e.dataTransfer.effectAllowed='move';if(node)e.dataTransfer.setDragImage(node,Math.min(node.offsetWidth/2,80),Math.min(node.offsetHeight/2,28));}}}}
function circled(k){{return {{A:'ⓐ',B:'ⓑ',C:'ⓒ',D:'ⓓ'}}[k]||k.toLowerCase()}}
function render(){{const wrap=el('screens');wrap.innerHTML='';QUIZ.questions.forEach((q,i)=>{{const sec=document.createElement('section');sec.className='screen qscreen';sec.id='q'+i;sec.style.backgroundImage=`linear-gradient(rgba(31,41,55,.24),rgba(31,41,55,.30)),linear-gradient(rgba(255,255,255,.42),rgba(245,243,255,.52)),url("${{bg}}")`;sec.innerHTML=`<div class="q-card"><div class="q-head"><div class="badge">Q${{i+1}}</div><div class="instruction">${{q.instruction}}</div><div class="tag">${{q.storyGrammar.replace('_',' ')}}</div></div><div id="body${{i}}"></div><div class="feedback" id="fb${{i}}"></div><div class="nav-row"><button class="btn btn-retry show" onclick="retry(${{i}})">Retry</button><button class="btn btn-check" id="check${{i}}" onclick="check(${{i}})">Check</button><button class="btn btn-next" id="next${{i}}" onclick="${{i===QUIZ.questions.length-1?'showStudent()':'go('+(i+1)+')'}}">${{i===QUIZ.questions.length-1?'See Results':'Next'}}</button></div></div>`;wrap.appendChild(sec);renderBody(q,i);}});updateNav();}}
function renderBody(q,i){{const b=el('body'+i);if(q.type==='story_sequence_drag'){{b.innerHTML=`<div class="slots seq-slots" id="slots${{i}}">${{q.interaction.correct.map((_,idx)=>`<div class="slot" ondragover="allowDrop(event)" ondrop="dropSeq(event,${{i}},${{idx}})" onclick="unplaceSeq(${{i}},${{idx}})">Scene ${{idx+1}}</div>`).join('')}}</div><div class="bank seq-bank" id="bank${{i}}"></div>`;q.resources.images.forEach(r=>addSeqCard(i,r));answers[i]=Array(q.interaction.correct.length).fill(null);}}
else if(q.type==='setting_slot_drag'){{const first=q.resources.images[0];b.innerHTML=`<div class="scene-grid single-grid"><div class="scene-card image-only"><img src="${{img(first.path)}}"></div></div><div class="slots">${{q.interaction.slots.map(s=>`<div class="slot" data-slot="${{s.key}}" ondragover="allowDrop(event)" ondrop="dropSetting(event,${{i}},'${{s.key}}')" onclick="placeSetting(${{i}},'${{s.key}}')">${{s.label}}</div>`).join('')}}</div><div class="bank setting-bank" id="bank${{i}}"></div>`;q.interaction.items.forEach(item=>addMiniCard(i,item));answers[i]={{}};}}
else if(q.type==='listen_scene_mcq'){{b.innerHTML=`<button class="audio-btn" onclick="new Audio('${{audio(q.resources.audio.path)}}').play()">▶</button><div class="scene-grid event-grid">${{q.interaction.options.map(o=>{{const r=q.resources.images.find(x=>x.sceneId===o.scene)||q.resources.images[0];return `<div class="scene-card image-only" data-opt="${{o.key}}" onclick="selectOpt(${{i}},'${{o.key}}')"><img src="${{img(r.path)}}"></div>`}}).join('')}}</div>`;}}
else if(q.type==='scene_word_unscramble'){{const first=q.resources.images[0];b.innerHTML=`<div class="scene-grid single-grid"><div class="scene-card image-only"><img src="${{img(first.path)}}"></div></div><div class="answer-row" id="answer${{i}}" ondragover="allowDrop(event)" ondrop="dropWord(event,${{i}})">Drop words here.</div><div class="word-bank" id="bank${{i}}"></div>`;q.interaction.items.forEach(w=>addWord(i,w));answers[i]=[];}}
else{{const first=q.resources.images[0];b.innerHTML=`<div class="scene-grid single-grid"><div class="scene-card image-only"><img src="${{img(first.path)}}"></div></div><div class="option-grid">${{q.interaction.options.map(o=>`<button class="option" data-opt="${{o.key}}" onclick="selectOpt(${{i}},'${{o.key}}')"><span class="opt-letter">${{circled(o.key)}}</span><span>${{o.text}}</span></button>`).join('')}}</div>`;}}}}
function addSeqCard(i,r){{const card=document.createElement('div');card.className='scene-card image-only';card.dataset.scene=r.sceneId;card.draggable=true;card.innerHTML=`<img src="${{img(r.path)}}">`;card.ondragstart=(e)=>dragStart(e,'seq',i,r.sceneId,card);card.onclick=()=>placeSeq(i,r.sceneId);el('bank'+i).appendChild(card)}}
function placeSeq(i,scene,pos=null){{if(scores[i]!=null)return;const arr=answers[i];if(pos===null)pos=arr.indexOf(null);if(pos<0)return;const old=arr.indexOf(scene);if(old>=0)arr[old]=null;arr[pos]=scene;paintSeq(i)}}
function dropSeq(e,i,idx){{allowDrop(e);if(dragged&&dragged.kind==='seq'&&dragged.i===i)placeSeq(i,dragged.value,idx);dragged=null}}
function unplaceSeq(i,idx){{if(scores[i]!=null)return;answers[i][idx]=null;paintSeq(i)}}
function paintSeq(i){{document.querySelectorAll(`#slots${{i}} .slot`).forEach((s,idx)=>{{const sc=answers[i][idx];s.innerHTML=sc?`<img src="${{img(QUIZ.questions[i].resources.images.find(r=>r.sceneId===sc).path)}}">`:`Scene ${{idx+1}}`;s.classList.toggle('filled',!!sc)}});document.querySelectorAll(`#bank${{i}} .scene-card`).forEach(c=>c.style.display=answers[i].includes(c.dataset.scene)?'none':'');}}
function addMiniCard(i,item){{const c=document.createElement('button');c.className='mini-card';c.textContent=item.text;c.dataset.key=item.key;c.draggable=true;c.ondragstart=(e)=>dragStart(e,'setting',i,item.key,c);c.onclick=()=>{{if(scores[i]!=null)return;selected=item.key;document.querySelectorAll(`#bank${{i}} .mini-card`).forEach(x=>x.classList.remove('selected'));c.classList.add('selected')}};el('bank'+i).appendChild(c)}}
function placeSetting(i,slot){{if(scores[i]!=null||!selected)return;const prev=answers[i][slot];if(prev){{const old=document.querySelector(`#bank${{i}} .mini-card[data-key="${{prev}}"]`);if(old)old.style.display='';}}answers[i][slot]=selected;const s=document.querySelector(`#body${{i}} .slot[data-slot="${{slot}}"]`);const item=QUIZ.questions[i].interaction.items.find(x=>x.key===selected);s.textContent=item.text;s.classList.add('filled');const used=document.querySelector(`#bank${{i}} .mini-card[data-key="${{selected}}"]`);if(used)used.style.display='none';selected=null;document.querySelectorAll(`#bank${{i}} .mini-card`).forEach(x=>x.classList.remove('selected'))}}
function dropSetting(e,i,slot){{allowDrop(e);if(dragged&&dragged.kind==='setting'&&dragged.i===i){{selected=dragged.value;placeSetting(i,slot)}}dragged=null}}
function addWord(i,w){{const c=document.createElement('button');c.className='word';c.textContent=w;c.draggable=true;c.ondragstart=(e)=>dragStart(e,'word',i,w,c);c.onclick=()=>addWordAnswer(i,w,c);el('bank'+i).appendChild(c)}}
function addWordAnswer(i,w,source=null){{if(scores[i]!=null)return;answers[i].push(w);if(source)source.style.display='none';else{{const btn=[...document.querySelectorAll(`#bank${{i}} .word`)].find(x=>x.textContent===w&&x.style.display!=='none');if(btn)btn.style.display='none';}}paintWords(i)}}
function dropWord(e,i){{allowDrop(e);if(dragged&&dragged.kind==='word'&&dragged.i===i)addWordAnswer(i,dragged.value);dragged=null}}
function removeWord(i,idx,w){{if(scores[i]!=null)return;answers[i].splice(idx,1);const btn=[...document.querySelectorAll(`#bank${{i}} .word`)].find(x=>x.textContent===w&&x.style.display==='none');if(btn)btn.style.display='';paintWords(i)}}
function paintWords(i){{const a=el('answer'+i);a.innerHTML=answers[i].length?answers[i].map((w,idx)=>`<button class="word" onclick="removeWord(${{i}},${{idx}},'${{w.replace(/'/g,"\\\\'")}}')">${{w}}</button>`).join(''):'Drop words here.'}}
function selectOpt(i,key){{if(scores[i]!=null)return;answers[i]=key;document.querySelectorAll(`#body${{i}} [data-opt]`).forEach(o=>o.classList.toggle('selected',o.dataset.opt===key))}}
function lockQuestion(i){{document.querySelectorAll(`#body${{i}} button,#body${{i}} .scene-card,#body${{i}} .slot`).forEach(n=>{{n.disabled=true;n.draggable=false;n.style.pointerEvents='none'}})}}
function check(i){{if(scores[i]!=null)return;const q=QUIZ.questions[i];let score=0;if(q.type==='story_sequence_drag'){{let total=0,got=0;q.scoring.components.forEach(c=>{{total+=Number(c.weight);const placed=answers[i].indexOf(c.key);const dist=placed<0?99:Math.abs(placed+1-Number(c.correctValue));got+=Number(c.weight)*Math.max(0,1-dist*.5)}});score=Math.round(got/total*100);answerDetails[i]={{score,order:[...answers[i]]}};}}
else if(q.type==='setting_slot_drag'){{let total=0,got=0;q.scoring.components.forEach(c=>{{total+=Number(c.weight);const placed=answers[i][c.key];const item=q.interaction.items.find(x=>x.key===placed);if(placed===c.correctValue)got+=Number(c.weight);else if(item&&item.slot===c.key)got+=Number(c.weight)*.35;}});score=Math.round(got/total*100);answerDetails[i]={{score,slots:{{...answers[i]}}}};}}
else if(q.type==='scene_word_unscramble'){{let total=0,got=0;q.scoring.components.forEach(c=>{{total+=Number(c.weight);if(answers[i][Number(c.correctValue)-1]===c.key)got+=Number(c.weight)}});score=Math.round(got/total*100);document.querySelectorAll(`#answer${{i}} .word`).forEach((n,idx)=>{{n.classList.add(answers[i][idx]===q.interaction.correct[idx]?'correct':'wrong')}});answerDetails[i]={{score,words:[...answers[i]]}};}}
else{{const o=q.interaction.options.find(x=>x.key===answers[i]);score=o?Number(o.score):0;document.querySelectorAll(`#body${{i}} [data-opt]`).forEach(n=>{{const opt=q.interaction.options.find(x=>x.key===n.dataset.opt);if(opt&&opt.isCorrect)n.classList.add('correct');else if(n.dataset.opt===answers[i])n.classList.add('wrong');}});answerDetails[i]={{score,selected:answers[i],selectedOption:o||null}};}}
scores[i]=score;lockQuestion(i);const fb=el('fb'+i);fb.className='feedback show '+(score>=85?'ok':'no');fb.textContent=(score>=85?'Great! ':'')+`Score: ${{score}} / 100`;el('next'+i).classList.add('show');updateNav();}}
function retry(i){{scores[i]=null;answers[i]=null;answerDetails[i]=null;renderBody(QUIZ.questions[i],i);el('fb'+i).className='feedback';el('next'+i).classList.remove('show');updateNav()}}
function showStudent(){{allScreens().forEach(s=>s.classList.remove('active'));el('student').classList.add('active');el('bookey').classList.remove('show');el('oxGrid').innerHTML=QUIZ.questions.map((q,i)=>`<div class="ox ${{scores[i]>=85?'ok':'no'}}">Q${{i+1}}<br>${{scores[i]>=85?'O':'X'}}</div>`).join('')}}
function sgScores(){{const out={{}};QUIZ.questions.forEach((q,i)=>out[q.storyGrammar]=scores[i]??0);return out}}
function parentComment(q,i){{const score=scores[i]??0;const detail=answerDetails[i]||{{}};if(score>=85)return `${{sgNames[q.storyGrammar]}} 영역을 안정적으로 이해했습니다. 이 문항의 핵심 단서와 정답 근거를 잘 연결했습니다.`;if(q.type==='story_sequence_drag'){{return score>=70?'전체 사건 흐름은 대체로 잡았지만, 일부 중간 장면의 앞뒤 관계가 흔들렸습니다. 사건의 원인-행동-결과 순서를 다시 말해보면 좋습니다.':'이야기의 결과와 사건 전개 순서를 연결하는 연습이 필요합니다. 처음 장면과 마지막 장면을 먼저 고정한 뒤 중간 사건을 배열해보세요.';}}
if(q.type==='setting_slot_drag'){{const gaps=q.scoring.components.filter(c=>(detail.slots||{{}})[c.key]!==c.correctValue).map(c=>c.key.replace('_',' '));return gaps.length?`Setting에서 ${{gaps.join(', ')}} 단서가 약했습니다. 첫 장면에서 인물, 장소, 처음 상황을 나누어 찾는 연습이 필요합니다.`:'Setting 단서를 대체로 이해했지만 일부 카드 범주가 흔들렸습니다.';}}
if(q.type==='scene_word_unscramble'){{return score>=70?'등장인물의 행동 문장을 대체로 이해했지만, 핵심 행동어의 위치가 일부 흔들렸습니다. 누가 무엇을 했는지 순서대로 말해보면 좋습니다.':'Attempt 문장에서 행동의 순서를 구성하는 데 어려움이 보였습니다. 이미지 속 행동을 먼저 말한 뒤 문장으로 배열하는 연습이 필요합니다.';}}
const opt=detail.selectedOption;if(opt){{if(opt.isCorrect)return `${{sgNames[q.storyGrammar]}} 문항을 정확히 이해했습니다.`;return `${{opt.score}}점 선택지입니다. ${{opt.diagnostic||'선택한 답이 문항의 핵심 단서와 충분히 연결되지 않았습니다.'}}`;}}
return `${{sgNames[q.storyGrammar]}} 영역에서 추가 확인이 필요합니다.`;}}
function showParent(){{allScreens().forEach(s=>s.classList.remove('active'));el('parent').classList.add('active');const s=sgScores();el('scoreList').innerHTML=QUIZ.questions.map((q,i)=>`<div class="score-card"><strong>Q${{i+1}} · ${{sgNames[q.storyGrammar]}}</strong><div class="score">${{scores[i]??0}} / 100</div><p>${{parentComment(q,i)}}</p></div>`).join('');drawRadar(s)}}
function drawRadar(s){{const c=el('radar'),ctx=c.getContext('2d'),cx=180,cy=150,r=105;ctx.clearRect(0,0,360,300);ctx.strokeStyle='#DDD6FE';ctx.fillStyle='#7C3AED';ctx.font='12px Arial';for(let level=1;level<=4;level++){{ctx.beginPath();sgOrder.forEach((k,i)=>{{const a=-Math.PI/2+i*Math.PI*2/6;const rr=r*level/4;const x=cx+Math.cos(a)*rr,y=cy+Math.sin(a)*rr;i?ctx.lineTo(x,y):ctx.moveTo(x,y)}});ctx.closePath();ctx.stroke();}}ctx.beginPath();sgOrder.forEach((k,i)=>{{const a=-Math.PI/2+i*Math.PI*2/6;const rr=r*((s[k]??0)/100);const x=cx+Math.cos(a)*rr,y=cy+Math.sin(a)*rr;i?ctx.lineTo(x,y):ctx.moveTo(x,y)}});ctx.closePath();ctx.fillStyle='rgba(124,58,237,.35)';ctx.fill();ctx.strokeStyle='#7C3AED';ctx.stroke();sgOrder.forEach((k,i)=>{{const a=-Math.PI/2+i*Math.PI*2/6;ctx.fillStyle='#374151';ctx.fillText(k.split('_')[0],cx+Math.cos(a)*(r+28)-24,cy+Math.sin(a)*(r+28)+4)}})}}
function restart(){{scores.fill(null);answers.fill(null);answerDetails.fill(null);allScreens().forEach(s=>s.classList.remove('active'));el('cover').classList.add('active');updateNav()}}
render();
</script>
</body>
</html>
"""


def build_html(quiz: dict[str, Any], out_path: Path) -> None:
    html = HTML_TEMPLATE.format(
        title=quiz["story"]["title"],
        story_id=quiz["story"]["storyId"],
        cover=quiz["assets"].get("coverImage", ""),
        bookey=quiz["assets"].get("hintCharacter", ""),
        quiz_json=json.dumps(quiz, ensure_ascii=False),
    )
    out_path.write_text(html, encoding="utf-8")


STORIES = [
    StoryDef(
        code="OG0021",
        title="Milo and the Lost Color",
        resource_dir="OG0021_Milo and the Lost Color",
        level="Level 2",
        sequence=["SC02", "SC03", "SC06", "SC08", "SC10"],
        sequence_titles={"SC02": "Milo wakes up gray", "SC03": "Milo enters the forest", "SC06": "Milo feels sad", "SC08": "Milo understands", "SC10": "Milo returns home"},
        setting={
            "who": {"correct": "milo", "items": [{"key": "milo", "text": "Milo"}, {"key": "butterfly", "text": "the butterfly", "diagnostic": "초반 주인공과 이후 만나는 인물을 혼동함"}]},
            "where": {"correct": "forest", "items": [{"key": "forest", "text": "in the forest"}, {"key": "blue_pond", "text": "by a blue pond", "diagnostic": "중간 장면의 장소를 처음 배경으로 혼동함"}]},
            "at_first": {"correct": "changes_colors", "items": [{"key": "changes_colors", "text": "loves to change colors"}, {"key": "looks_mirror", "text": "looks in the mirror", "diagnostic": "결말 장면을 처음 상황으로 혼동함"}]},
        },
        event_scene="SC02",
        event_sentence="SC02_ST01_N",
        event_options=[
            {"key": "A", "scene": "SC02", "text": "Milo wakes up gray.", "score": 100, "isCorrect": True},
            {"key": "B", "scene": "SC04", "text": "The butterfly keeps its color.", "score": 30, "diagnostic": "문제의 시작이 아니라 첫 만남 장면을 선택함"},
            {"key": "C", "scene": "SC07", "text": "The pond shines.", "score": 20, "diagnostic": "해결의 실마리 장면을 사건 시작으로 혼동함"},
            {"key": "D", "scene": "SC10", "text": "Milo looks in the mirror.", "score": 0, "diagnostic": "결말 장면을 사건 시작으로 혼동함"},
        ],
        attempt_scene="SC03",
        attempt_sentence="Milo walks into the forest.",
        reaction_scene="SC06",
        reaction_options=[
            {"key": "A", "text": "sad", "score": 100, "isCorrect": True},
            {"key": "B", "text": "happy", "score": 0, "diagnostic": "해결 이후 감정을 슬픈 장면에 적용함"},
            {"key": "C", "text": "angry", "score": 40, "diagnostic": "부정 감정은 파악했지만 슬픔과 분노를 혼동함"},
            {"key": "D", "text": "proud", "score": 10, "diagnostic": "결말의 자신감을 현재 감정으로 혼동함"},
        ],
        internal_scene="SC08",
        internal_options=[
            {"key": "A", "text": "Maybe my color is inside me!", "score": 100, "isCorrect": True},
            {"key": "B", "text": "I need the butterfly's color.", "score": 25, "diagnostic": "외부에서 색을 얻으려는 초반 생각에 머무름"},
            {"key": "C", "text": "I should stop looking.", "score": 20, "diagnostic": "슬픈 반응과 깨달음을 혼동함"},
            {"key": "D", "text": "The pond is my home.", "score": 0, "diagnostic": "장면의 배경 단서를 내적 깨달음으로 오해함"},
        ],
        hints=["Milo loses his color, looks for it, and finds it inside.", "Who is there? Where is he?", "Listen for Milo's problem.", "Put the short action sentence in order.", "Milo is by the pond. How does he feel?", "Think about what Milo learns."],
        setting_scene="SC01",
    ),
    StoryDef(
        code="OG0036",
        title="The Secret of Judy's Silver Racket",
        resource_dir="OG0036_The Secret of Judy's Silver Racket",
        level="Level 3",
        sequence=["SC01", "SC02", "SC08", "SC12", "SC16"],
        sequence_titles={"SC01": "Judy is nervous", "SC02": "Judy gets a silver racket", "SC08": "The racket breaks", "SC12": "Judy realizes the truth", "SC16": "Judy finds confidence"},
        setting={
            "who": {"correct": "judy", "items": [{"key": "judy", "text": "Judy"}, {"key": "bella", "text": "Bella", "diagnostic": "후반 상대 선수를 처음 주인공으로 혼동함"}]},
            "where": {"correct": "tennis_court", "items": [{"key": "tennis_court", "text": "on the tennis court"}, {"key": "shop", "text": "at Jingle Jump", "diagnostic": "문제 이후의 장소를 처음 배경으로 혼동함"}]},
            "at_first": {"correct": "nervous", "items": [{"key": "nervous", "text": "feels nervous in matches"}, {"key": "confident", "text": "finds her confidence", "diagnostic": "결말의 변화를 처음 상황으로 혼동함"}]},
        },
        event_scene="SC01",
        event_sentence="SC01_ST03_N",
        event_options=[
            {"key": "A", "scene": "SC01", "text": "Judy loses because of fear.", "score": 100, "isCorrect": True},
            {"key": "B", "scene": "SC02", "text": "Judy gets a silver racket.", "score": 35, "diagnostic": "문제 해결 도구의 등장을 문제의 시작으로 혼동함"},
            {"key": "C", "scene": "SC08", "text": "The racket breaks.", "score": 25, "diagnostic": "중간 위기를 처음 문제로 혼동함"},
            {"key": "D", "scene": "SC16", "text": "Judy finds confidence.", "score": 0, "diagnostic": "결말 장면을 사건 시작으로 혼동함"},
        ],
        attempt_scene="SC13",
        attempt_sentence="She stood up and held her old racket tightly.",
        reaction_scene="SC09",
        reaction_options=[
            {"key": "A", "text": "scared", "score": 100, "isCorrect": True},
            {"key": "B", "text": "happy", "score": 0, "diagnostic": "초반 승리 감정을 위기 장면에 적용함"},
            {"key": "C", "text": "angry", "score": 35, "diagnostic": "부정 감정은 파악했지만 두려움과 분노를 혼동함"},
            {"key": "D", "text": "proud", "score": 15, "diagnostic": "결말의 자신감을 위기 장면에 투영함"},
        ],
        internal_scene="SC12",
        internal_options=[
            {"key": "A", "text": "The whispers came from Judy herself.", "score": 100, "isCorrect": True},
            {"key": "B", "text": "The silver racket is magic.", "score": 20, "diagnostic": "겉으로 보이는 도구를 내적 능력보다 우선함"},
            {"key": "C", "text": "Bella is too strong.", "score": 25, "diagnostic": "상대 선수에 대한 부담과 깨달음을 혼동함"},
            {"key": "D", "text": "She should stop playing.", "score": 0, "diagnostic": "포기 반응을 내적 성장으로 오해함"},
        ],
        hints=["Judy trusts a racket, then learns to trust herself.", "Who is there? Where is she?", "Listen for Judy's problem.", "Build the sentence about Judy's brave action.", "The whispers stop. How does Judy feel?", "Think about what Judy learns about herself."],
    ),
    StoryDef(
        code="OG0049",
        title="The Mystery of the Deep: The Lost Light",
        resource_dir="OG0049_The Mystery of the Deep The Lost Light",
        level="Level 3",
        sequence=["SC01", "SC06", "SC09", "SC12", "SC16"],
        sequence_titles={"SC01": "The deep ocean is dark", "SC06": "A bag covers the fish", "SC09": "Kira and Toby act", "SC12": "The bag comes off", "SC16": "The light guides the way"},
        setting={
            "who": {"correct": "kira_toby", "items": [{"key": "kira_toby", "text": "Kira and Toby"}, {"key": "anglerfish", "text": "the anglerfish", "diagnostic": "도움을 받는 대상을 초반 탐사 인물로 혼동함"}]},
            "where": {"correct": "deep_ocean", "items": [{"key": "deep_ocean", "text": "deep in the ocean"}, {"key": "jagged_cave", "text": "near a jagged cave", "diagnostic": "위기 장소를 이야기의 처음 배경으로 혼동함"}]},
            "at_first": {"correct": "see_light", "items": [{"key": "see_light", "text": "see a tiny golden light"}, {"key": "remove_bag", "text": "remove the plastic bag", "diagnostic": "해결 행동을 처음 상황으로 혼동함"}]},
        },
        event_scene="SC06",
        event_sentence="SC06_ST01_N",
        event_options=[
            {"key": "A", "scene": "SC02", "text": "Kira sees a tiny golden light.", "score": 30, "diagnostic": "문제의 단서와 문제 자체를 혼동함"},
            {"key": "B", "scene": "SC06", "text": "A plastic bag covers the fish.", "score": 100, "isCorrect": True},
            {"key": "C", "scene": "SC09", "text": "Kira and Toby decide to help.", "score": 25, "diagnostic": "문제 이후의 대응 장면을 사건 시작으로 혼동함"},
            {"key": "D", "scene": "SC13", "text": "The fish's lamp shines again.", "score": 0, "diagnostic": "해결 장면을 사건 시작으로 혼동함"},
        ],
        attempt_scene="SC09",
        attempt_sentence="Diver Toby opened the submarine's silver robot arms.",
        reaction_scene="SC08",
        reaction_options=[
            {"key": "A", "text": "afraid", "score": 100, "isCorrect": True},
            {"key": "B", "text": "happy", "score": 0, "diagnostic": "해결 이후 감정을 위기 장면에 적용함"},
            {"key": "C", "text": "angry", "score": 30, "diagnostic": "위기감은 파악했지만 두려움과 분노를 혼동함"},
            {"key": "D", "text": "proud", "score": 10, "diagnostic": "구조 이후의 감정을 현재 장면에 투영함"},
        ],
        internal_scene="SC09",
        internal_options=[
            {"key": "A", "text": "We must act now!", "score": 100, "isCorrect": True},
            {"key": "B", "text": "The light is not important.", "score": 0, "diagnostic": "중심 문제를 놓치고 반대되는 생각을 선택함"},
            {"key": "C", "text": "The fish can fix it alone.", "score": 20, "diagnostic": "도움이 필요한 상황과 인물의 의도를 연결하지 못함"},
            {"key": "D", "text": "The cave is beautiful.", "score": 10, "diagnostic": "위험 단서를 배경 묘사로만 이해함"},
        ],
        hints=["A deep-sea fish loses her light, and the team helps her.", "Who is there? Where does the story start?", "Listen for the real problem.", "Build the sentence about Toby's rescue action.", "The fish is near danger. How does she feel?", "Think about what Toby wants to do now."],
        setting_scene="SC03",
    ),
    StoryDef(
        code="CS0003",
        title="Hans in Luck",
        resource_dir="CS0003_Hans in Luck",
        level="Level 2",
        sequence=["SC01", "SC03", "SC07", "SC13", "SC14"],
        sequence_titles={"SC01": "Hans gets gold", "SC03": "Gold for horse", "SC07": "Cow for pig", "SC13": "Stone falls", "SC14": "Hans feels free"},
        setting={
            "who": {"correct": "hans", "items": [{"key": "hans", "text": "Hans"}, {"key": "horse_man", "text": "a man on a horse", "diagnostic": "첫 장면의 주인공과 이후 만나는 인물을 혼동함"}]},
            "where": {"correct": "road_home", "items": [{"key": "road_home", "text": "on the road home"}, {"key": "well", "text": "at a well", "diagnostic": "마지막 장면의 장소를 시작 배경으로 혼동함"}]},
            "at_first": {"correct": "receives_gold", "items": [{"key": "receives_gold", "text": "receives a big piece of gold"}, {"key": "loses_stone", "text": "loses the stone", "diagnostic": "이야기의 결과를 처음 상황으로 혼동함"}]},
        },
        event_scene="SC02",
        event_sentence="SC02_ST01_N",
        event_options=[
            {"key": "A", "scene": "SC04", "text": "Hans rides the horse.", "score": 25, "diagnostic": "문제 발생 뒤의 이동 장면을 사건 시작으로 혼동함"},
            {"key": "B", "scene": "SC02", "text": "The gold is too heavy.", "score": 100, "isCorrect": True},
            {"key": "C", "scene": "SC07", "text": "Hans trades the cow.", "score": 20, "diagnostic": "반복되는 거래 장면을 사건 시작으로 혼동함"},
            {"key": "D", "scene": "SC14", "text": "Hans feels free.", "score": 0, "diagnostic": "결말 장면을 사건 시작으로 혼동함"},
        ],
        attempt_scene="SC09",
        attempt_sentence="Take this pig and give me your goose.",
        reaction_scene="SC09",
        reaction_options=[
            {"key": "A", "text": "happy", "score": 20, "diagnostic": "결말의 기쁨을 문제 장면에 적용함"},
            {"key": "B", "text": "afraid", "score": 100, "isCorrect": True},
            {"key": "C", "text": "angry", "score": 35, "diagnostic": "부정 감정은 파악했으나 두려움과 분노를 혼동함"},
            {"key": "D", "text": "proud", "score": 0, "diagnostic": "장면의 위험 신호를 놓침"},
        ],
        internal_scene="SC14",
        internal_options=[
            {"key": "A", "text": "I am free now.", "score": 100, "isCorrect": True},
            {"key": "B", "text": "I need more gold.", "score": 10, "diagnostic": "초반 목표를 결말의 깨달음으로 혼동함"},
            {"key": "C", "text": "I want the horse back.", "score": 20, "diagnostic": "이전 거래 장면에 머무름"},
            {"key": "D", "text": "I am in trouble.", "score": 30, "diagnostic": "중간 위기 장면의 생각을 결말에 적용함"},
        ],
        hints=["Hans trades many things. Think from gold to the stone.", "Who is there? Where is he?", "Listen for the first problem.", "Start with the action word in the sentence.", "Look at Hans's face and the danger.", "Think about why Hans is happy at the end."],
    ),
    StoryDef(
        code="CS0006",
        title="The Cat in Boots",
        resource_dir="CS0006_The Cat in Boots",
        level="Level 2",
        sequence=["SC01", "SC02", "SC10", "SC18", "SC20"],
        sequence_titles={"SC01": "Young man gets a cat", "SC02": "Cat promises help", "SC10": "Cat tricks the King", "SC18": "Cat beats the Giant", "SC20": "Happy ending"},
        setting={
            "who": {"correct": "youngest_son", "items": [{"key": "youngest_son", "text": "the youngest son"}, {"key": "king", "text": "the King", "diagnostic": "뒤에 등장하는 인물을 시작 장면의 중심 인물로 혼동함"}]},
            "where": {"correct": "mill_home", "items": [{"key": "mill_home", "text": "at the mill"}, {"key": "castle", "text": "at the castle", "diagnostic": "후반 배경을 처음 배경으로 혼동함"}]},
            "at_first": {"correct": "gets_cat", "items": [{"key": "gets_cat", "text": "gets only a cat"}, {"key": "marries_princess", "text": "marries the Princess", "diagnostic": "결말을 처음 상황으로 혼동함"}]},
        },
        event_scene="SC01",
        event_sentence="SC01_ST02_N",
        event_options=[
            {"key": "A", "scene": "SC01", "text": "The youngest son gets only a cat.", "score": 100, "isCorrect": True},
            {"key": "B", "scene": "SC05", "text": "The Cat gives a rabbit to the King.", "score": 20, "diagnostic": "문제 이후의 첫 성과 장면을 사건 시작으로 혼동함"},
            {"key": "C", "scene": "SC10", "text": "The Cat cries for help.", "score": 25, "diagnostic": "중간 계획 장면을 시작 사건으로 혼동함"},
            {"key": "D", "scene": "SC18", "text": "The Cat catches the mouse.", "score": 10, "diagnostic": "해결 장면을 시작 사건으로 혼동함"},
        ],
        attempt_scene="SC18",
        attempt_sentence="The Cat jumped on the mouse.",
        reaction_scene="SC16",
        reaction_options=[
            {"key": "A", "text": "afraid", "score": 100, "isCorrect": True},
            {"key": "B", "text": "happy", "score": 20, "diagnostic": "장면의 위험보다 성공 결과를 먼저 떠올림"},
            {"key": "C", "text": "angry", "score": 35, "diagnostic": "긴장감은 파악했으나 감정을 혼동함"},
            {"key": "D", "text": "bored", "score": 0, "diagnostic": "장면의 긴박함을 놓침"},
        ],
        internal_scene="SC17",
        internal_options=[
            {"key": "A", "text": "I can trick the Giant.", "score": 100, "isCorrect": True},
            {"key": "B", "text": "I should run away now.", "score": 30, "diagnostic": "두려움 반응과 다음 계획을 혼동함"},
            {"key": "C", "text": "I want to give the King a rabbit.", "score": 15, "diagnostic": "초반 행동을 후반 생각으로 혼동함"},
            {"key": "D", "text": "The castle is too small.", "score": 0, "diagnostic": "이야기와 무관한 생각을 선택함"},
        ],
        hints=["The Cat makes a poor son look rich. Think of the big steps.", "Who is there? Where does the story start?", "Listen for the youngest son's problem.", "Make the short action sentence.", "Look at the Giant. How does the Cat feel?", "Think about the Cat's clever plan."],
    ),
    StoryDef(
        code="OG0005",
        title="The Rainbow Cloud in the Box",
        resource_dir="OG0005_The Rainbow Cloud in the Box",
        level="Level 4",
        sequence=["SC02", "SC03", "SC05", "SC07", "SC12"],
        sequence_titles={"SC02": "Rainbow cloud appears", "SC03": "Didi wants it", "SC05": "Didi catches it", "SC07": "Cloud turns gray", "SC12": "Cloud comes back"},
        setting={
            "who": {"correct": "podo_didi", "items": [{"key": "podo_didi", "text": "Podo and Didi"}, {"key": "cloud", "text": "the rainbow cloud", "diagnostic": "대상과 인물을 혼동함"}]},
            "where": {"correct": "tiny_rock", "items": [{"key": "tiny_rock", "text": "on Tiny Rock"}, {"key": "crystal_box", "text": "in a crystal box", "diagnostic": "중간 사건의 장소를 시작 배경으로 혼동함"}]},
            "at_first": {"correct": "watch_universe", "items": [{"key": "watch_universe", "text": "watch the universe"}, {"key": "capture_cloud", "text": "capture the cloud", "diagnostic": "시도 장면을 처음 상황으로 혼동함"}]},
        },
        event_scene="SC03",
        event_sentence="SC03_ST03_N",
        event_options=[
            {"key": "A", "scene": "SC01", "text": "Podo and Didi watch the sky.", "score": 30, "diagnostic": "배경 장면을 사건 시작으로 혼동함"},
            {"key": "B", "scene": "SC03", "text": "Didi wants to capture the cloud.", "score": 100, "isCorrect": True},
            {"key": "C", "scene": "SC07", "text": "The cloud turns gray.", "score": 20, "diagnostic": "문제의 시작이 아니라 결과 장면을 선택함"},
            {"key": "D", "scene": "SC12", "text": "The cloud comes back to life.", "score": 0, "diagnostic": "해결 장면을 발단으로 혼동함"},
        ],
        attempt_scene="SC10",
        attempt_sentence="He picked up the box and ran back outside.",
        reaction_scene="SC08",
        reaction_options=[
            {"key": "A", "text": "disappointed", "score": 100, "isCorrect": True},
            {"key": "B", "text": "proud", "score": 20, "diagnostic": "포획 직후 감정을 문제 장면에 적용함"},
            {"key": "C", "text": "angry", "score": 40, "diagnostic": "부정 감정은 파악했으나 실망과 분노를 혼동함"},
            {"key": "D", "text": "sleepy", "score": 0, "diagnostic": "장면의 문제 상황을 놓침"},
        ],
        internal_scene="SC10",
        internal_options=[
            {"key": "A", "text": "I made a mistake.", "score": 100, "isCorrect": True},
            {"key": "B", "text": "I need a bigger box.", "score": 20, "diagnostic": "문제의 원인을 소유가 아닌 도구로 오해함"},
            {"key": "C", "text": "The cloud is mine forever.", "score": 0, "diagnostic": "인물의 깨달음과 반대되는 생각을 선택함"},
            {"key": "D", "text": "Podo should go away.", "score": 10, "diagnostic": "친구의 조언을 내면 변화와 연결하지 못함"},
        ],
        hints=["Didi sees a cloud, traps it, and lets it go.", "Who is there? Where are they?", "Listen for Didi's wish. This starts the trouble.", "Start with He. Look for the action.", "The magic is gone. How does Didi feel?", "Think about what Didi learns."],
    ),
]


def build_all() -> None:
    for story in STORIES:
        src_dir = RESOURCES / story.resource_dir
        txt_path = next(src_dir.glob(f"{story.code}_*.txt"))
        _, scenes, emotions = parse_story(txt_path)
        out_dir = OUT_ROOT / story.code
        ensure_dirs(out_dir)
        assets = copy_assets(story, src_dir, out_dir)
        quiz = make_quiz(story, scenes, emotions, assets)
        build_reading_xlsx(quiz, out_dir / f"{story.code}_ReadingQuiz.xlsx")
        build_dev_xlsx(quiz, out_dir / f"{story.code}_DevSpec.xlsx")
        build_html(quiz, out_dir / f"{story.code}_ReadingQuiz.html")
        (out_dir / f"{story.code}.quiz.json").write_text(json.dumps(quiz, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Generated {story.code}: {out_dir}")


if __name__ == "__main__":
    build_all()
