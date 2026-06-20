"""
OG0005_ReadingQuiz.xlsx  —  exact OG0021 formatting clone
Colors, fonts, row heights, col widths, merges all match OG0021_ReadingQuiz.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette (exact OG0021 fills) ─────────────────────────
def P(hex6): return PatternFill('solid', fgColor='00'+hex6)

DARK_BLUE  = P('1F4E79')   # title row bg
MED_BLUE   = P('2E75B6')   # section banner / col header bg
LT_BLUE    = P('BDD7EE')   # TOTAL row / section-C note / matrix row-label bg / sub-header
YELLOW     = P('FFF2CC')   # question text / anchor scene rows / Q3 word rows / LRS-mapping Q rows
LT_GREEN   = P('E2EFDA')   # correct answer rows / LRS row bg / sg-section A col
GREEN      = P('70AD47')   # score=100 / max-pts col in Q3
GOLD       = P('FFD966')   # score 30-50 / partial credit medium
LT_ORANGE  = P('FFEECC')   # score 10-25
PINK_RED   = P('FCE4D6')   # Emerging band
LT_RED     = P('FFCCCC')   # score 0 / Beginning band

# ── Fonts ─────────────────────────────────────────────────
def FW(sz=10, bold=True):  return Font(bold=bold, color='FFFFFF', name='Calibri', size=sz)
def FB(sz=9,  bold=False): return Font(bold=bold, color='000000', name='Calibri', size=sz)
def FD(sz=9):              return Font(name='Calibri', size=sz)   # default (auto color)
def FG(sz=9):              return Font(name='Calibri', size=sz, color='555555', italic=True)

ALIGN_CW  = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LW  = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
ALIGN_C   = Alignment(horizontal='center', vertical='top',    wrap_text=None)
ALIGN_CWV = Alignment(horizontal='center', vertical='top',    wrap_text=True)

thin  = Side(style='thin', color='AAAAAA')
BORD  = Border(left=thin, right=thin, top=thin, bottom=thin)

def sc(ws, r, c, val, fill=None, font=None, align=ALIGN_LW):
    """Set cell value + style"""
    cell = ws.cell(r, c, val)
    if fill:  cell.fill  = fill
    if font:  cell.font  = font
    else:     cell.font  = FD()
    cell.alignment = align
    cell.border    = BORD
    return cell

def title(ws, r, c, val, fill=DARK_BLUE, sz=11):
    return sc(ws, r, c, val, fill=fill, font=FW(sz=sz), align=ALIGN_CW)

def hdr(ws, r, c, val, fill=MED_BLUE, fc='FFFFFF'):
    f = Font(bold=True, color=fc, name='Calibri', size=10)
    return sc(ws, r, c, val, fill=fill, font=f, align=ALIGN_CWV)

def banner(ws, r, c, val):
    return sc(ws, r, c, val, fill=MED_BLUE, font=FW(sz=10), align=ALIGN_CW)

def score_fill(s):
    """Color by score value — exact OG0021 scheme"""
    if s == 100: return GREEN
    if s >= 30:  return GOLD
    if s >= 10:  return LT_ORANGE
    return LT_RED

def set_cols(ws, widths_dict):
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w

def set_rows(ws, heights_dict):
    for row, h in heights_dict.items():
        ws.row_dimensions[row].height = h

def merge_title(ws, row, cols='J'):
    ws.merge_cells(f'A{row}:{cols}{row}')

# ══════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════
# QUIZ_LIST
# ══════════════════════════════════════════════════════════
ws = wb.active; ws.title = 'QUIZ_LIST'
set_cols(ws, {'A':8,'B':18,'C':22,'D':20,'E':14,'F':14,'G':16,'H':22,'I':22,'J':16})
set_rows(ws, {1:28,2:30,3:28,4:28,5:28,6:28,7:28,8:28,10:20})

title(ws, 1, 1, 'OG0005 Reading Quiz — Master List (Podo and Didi)', sz=12)
merge_title(ws, 1, 'J')

hdrs2 = ['Q_ID','Quiz Type','Story Grammar Element','Question (EN)',
         'Scene Ref','Answer Sheet','Max Score','LRS sg_element','Scoring Mode','Sheet Ref']
for c, h in enumerate(hdrs2, 1): hdr(ws, 2, c, h)

rows = [
    ['Q01','Scene Sequencing','Setting→IE→Attempt→Consequence→Resolution',
     'Put the 5 scenes in story order.',
     'SC02/SC03/SC05/SC07/SC12','SC02-SC03-SC05-SC07-SC12','100',
     'setting,initiating_event,attempt,consequence,resolution','Weighted Position','Q01_SEQUENCING'],
    ['Q02','Sentence-Scene Match','Initiating Event',
     'Listen and choose the matching scene.',
     'SC03/SC02/SC05/SC09','SC02','100','initiating_event','Weighted MCQ','Q02_SENT_MATCH'],
    ['Q03','Sentence Unscramble','Attempt',
     'Put the words in order to make a sentence.',
     '-','By trapping it, you extinguished its light.','100','attempt','Weighted Unscramble','Q03_UNSCRAMBLE'],
    ['Q04','Emotion Identification','Reaction / Internal Response',
     'How does Didi feel in this scene?',
     'SC08','Option B','100','reaction','Weighted MCQ','Q04_EMOTION'],
    ['Q05','Character Goal MCQ','Initiating Event / Goal',
     'What does Didi want at the start of the story?',
     '-','Option A','100','initiating_event','Weighted MCQ','Q05_CHAR_GOAL'],
    ['Q06','Main Idea MCQ','Theme / Resolution',
     'What lesson did Didi learn from this experience?',
     '-','Option B','100','theme','Weighted MCQ','Q06_MAIN_IDEA'],
]
for r, row in enumerate(rows, 3):
    for c, v in enumerate(row, 1):
        sc(ws, r, c, v, font=FD(9))

note = ws.cell(10, 1, 'Each question sheet defines per-option/position scores and weights. The HTML file reads this workbook via SheetJS to render the live quiz.')
note.font = FG(9); note.alignment = ALIGN_LW
ws.merge_cells('A10:J10')

# ══════════════════════════════════════════════════════════
# Q01_SEQUENCING
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('Q01_SEQUENCING')
set_cols(ws, {'A':8,'B':14,'C':30,'D':12,'E':12,'F':14,'G':12,'H':12,'I':20,'J':28})
set_rows(ws, {1:26,2:22,3:22,4:22,5:22,7:22,
              9:30,10:30,11:30,12:30,13:30,
              16:22,18:22,19:22,20:22,21:22,22:22,
              24:22,26:20,27:20,28:20,29:20,31:24})

title(ws, 1, 1, 'Q01 — Scene Sequencing | Story Grammar: Full Arc')
ws.merge_cells('A1:J1')
sc(ws, 2, 1, 'Question: Put the 5 scene images in the correct story order by dragging them into position.',
   fill=YELLOW, font=Font(bold=True, name='Calibri', size=9))
ws.merge_cells('A2:J2')
sc(ws, 3, 1, 'Scenes provided:  OG0005_SC02_I  |  OG0005_SC03_I  |  OG0005_SC05_I  |  OG0005_SC07_I  |  OG0005_SC12_I',
   fill=YELLOW)
ws.merge_cells('A3:J3')
sc(ws, 4, 1, 'Correct Answer: SC02 → SC03 → SC05 → SC07 → SC12',
   fill=LT_GREEN, font=Font(bold=True, name='Calibri', size=9))
ws.merge_cells('A4:J4')
sc(ws, 5, 1, 'Story Grammar: SC02=Initiating Event (cloud appears) | SC03=1st Observation (Didi notices) | SC05=Attempt (captures cloud) | SC07=Consequence (cloud turns gray) | SC12=Resolution (cloud freed)',
   fill=LT_BLUE)
ws.merge_cells('A5:J5')

banner(ws, 7, 1, 'SECTION A — Position Weight per Scene')
ws.merge_cells('A7:J7')

for c, h in enumerate(['Scene Asset','Correct Pos','Story Grammar Role','Position Weight','Weight Rationale','Max Points'], 1):
    hdr(ws, 8, c, h)

# SC02/SC12 are anchors (w=2.5) → YELLOW; SC03/SC05/SC07 mid (w=1.5) → LT_GREEN
seq_data = [
    ('OG0005_SC02_I', 1, 'Initiating Event (rainbow cloud appears)',    2.5,
     'Unambiguous opening event. Placing wrong = cannot identify story trigger.',
     '=D9/SUM($D$9:$D$13)*100',  YELLOW),
    ('OG0005_SC03_I', 2, '1st Observation (Didi notices cloud)',         1.5,
     'Clear observation scene; SC03/SC05 confusion forgivable — both mid-story.',
     '=D10/SUM($D$9:$D$13)*100', LT_GREEN),
    ('OG0005_SC05_I', 3, 'Attempt (Didi captures cloud)',                1.5,
     'Action scene similar to SC03; mixing these two penalized less.',
     '=D11/SUM($D$9:$D$13)*100', LT_GREEN),
    ('OG0005_SC07_I', 4, 'Consequence (cloud turns gray)',               1.5,
     'Consequence scene; close to SC05 thematically — smaller penalty for adjacency.',
     '=D12/SUM($D$9:$D$13)*100', LT_GREEN),
    ('OG0005_SC12_I', 5, 'Resolution (cloud freed, colors return)',      2.5,
     'Unambiguous final scene. Placing wrong = fundamental arc gap.',
     '=D13/SUM($D$9:$D$13)*100', YELLOW),
]
for i, (asset, pos, role, wt, rat, formula, row_fill) in enumerate(seq_data):
    r = 9 + i
    sc(ws, r, 1, asset,   fill=row_fill)
    sc(ws, r, 2, pos,     fill=row_fill, align=ALIGN_C)
    sc(ws, r, 3, role,    fill=row_fill)
    sc(ws, r, 4, wt,      fill=row_fill, align=ALIGN_C)
    sc(ws, r, 5, rat,     fill=row_fill)
    sc(ws, r, 6, formula, fill=LT_GREEN, align=ALIGN_C)

# TOTAL row
sc(ws, 14, 1, 'TOTAL',          fill=LT_BLUE, font=Font(bold=True,name='Calibri',size=9), align=ALIGN_C)
sc(ws, 14, 4, '=SUM(D9:D13)',   fill=LT_BLUE, font=Font(bold=True,name='Calibri',size=9), align=ALIGN_C)
sc(ws, 14, 6, '=SUM(F9:F13)',   fill=LT_BLUE, font=Font(bold=True,name='Calibri',size=9), align=ALIGN_C)

banner(ws, 16, 1, 'SECTION B — Score Matrix: Points per (Scene, Submitted Position)')
ws.merge_cells('A16:J16')

sc(ws, 17, 1, 'Scene \\ Submitted Pos', font=Font(bold=True,name='Calibri',size=9))
for c, h in enumerate(['Pos 1','Pos 2','Pos 3','Pos 4','Pos 5'], 2):
    hdr(ws, 17, c, h)

# weights: SC02=2.5, SC03=1.5, SC05=1.5, SC07=1.5, SC12=2.5  total=9.5
# score(placed,correct,w) = w * max(0, 1 - |placed-correct|*0.5) / 9.5 * 100
matrix = [
    ('OG0005_SC02_I', [26.3, 13.2,  0.0,  0.0,  0.0]),
    ('OG0005_SC03_I', [ 7.9, 15.8,  7.9,  0.0,  0.0]),
    ('OG0005_SC05_I', [ 0.0,  7.9, 15.8,  7.9,  0.0]),
    ('OG0005_SC07_I', [ 0.0,  0.0,  7.9, 15.8,  7.9]),
    ('OG0005_SC12_I', [ 0.0,  0.0,  0.0, 13.2, 26.3]),
]
for i, (label, vals) in enumerate(matrix):
    r = 18 + i
    sc(ws, r, 1, label, fill=LT_BLUE, font=Font(bold=True,name='Calibri',size=9), align=ALIGN_LW)
    for j, v in enumerate(vals, 2):
        if v > 20:   # diagonal (max)
            cell_fill = GREEN
            cell_font = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
        elif v > 0:  # partial credit
            cell_fill = GOLD
            cell_font = FB(9)
        else:        # zero
            cell_fill = LT_RED
            cell_font = FB(9)
        sc(ws, r, j, v, fill=cell_fill, font=cell_font, align=ALIGN_C)

banner(ws, 24, 1, 'SECTION C — Example Answer Scores')
ws.merge_cells('A24:J24')

for c, h in enumerate(['Example Student Answer','Submitted Order','Score','Interpretation'], 1):
    hdr(ws, 25, c, h, fill=LT_BLUE, fc='000000')

examples = [
    ('Perfect',             'SC02-SC03-SC05-SC07-SC12', 100, 'Full story arc understood'),
    ('SC05/SC07 swapped',   'SC02-SC03-SC07-SC05-SC12',  79, 'Adjacent attempt/consequence swap — mild penalty'),
    ('Middle scrambled',    'SC02-SC07-SC05-SC03-SC12',  53, 'Anchor scenes correct; internal sequence weak'),
    ('Beginning/end wrong', 'SC03-SC02-SC12-SC07-SC05',  28, 'Story arc fundamentally misunderstood'),
]
for i, (label, order, score, interp) in enumerate(examples):
    r = 26 + i
    sc(ws, r, 1, label, font=FD(9))
    sc(ws, r, 2, order, font=FD(9))
    sc(ws, r, 3, score, fill=score_fill(score), font=FD(9))
    sc(ws, r, 4, interp, font=FD(9))

lrs = ws.cell(31, 1, 'LRS: verb="answered" | object="quiz_OG0005_Q01_sequencing" | result.sg_element="setting,initiating_event,attempt,consequence,resolution" | result.sequence_submitted=[] | result.score_raw=<n> | result.position_scores={}')
lrs.fill = LT_GREEN; lrs.font = FW(sz=10); lrs.alignment = ALIGN_CW; lrs.border = BORD
ws.merge_cells('A31:J31')

# ══════════════════════════════════════════════════════════
# Q02_SENT_MATCH
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('Q02_SENT_MATCH')
set_cols(ws, {'A':8,'B':20,'C':36,'D':12,'E':20,'F':14,'G':24,'H':24,'I':22})
set_rows(ws, {1:26,5:30,6:36,7:36,8:36,9:36,11:22})

title(ws, 1, 1, 'Q02 — Sentence-Scene Match | Story Grammar: Initiating Event')
ws.merge_cells('A1:I1')
sc(ws, 2, 1, 'Question: Listen to the sentence. Which scene image does this sentence come from?',
   fill=YELLOW, font=Font(bold=True, name='Calibri', size=9))
ws.merge_cells('A2:I2')
sc(ws, 3, 1, 'Target Sentence (SC02_ST01_N): "But look! A rainbow stardust cloud is drifting close to Tiny Rock!"',
   fill=YELLOW)
ws.merge_cells('A3:I3')

for c, h in enumerate(['Option','Scene Image','Scene Description','Correct?','Score',
                        'Story Grammar Role','Distractor Rationale','Weakness Signal','LRS sg_element'], 1):
    hdr(ws, 5, c, h)

# option rows: col A (option letter), D (Correct?), E (Score) colored by score
# B,C,F,G,H,I: plain (no fill for non-answer cols — match OG0021)
q2_data = [
    ('A', 'OG0005_SC03_I', 'Didi observes the cloud closely',              'NO',  25,
     'Attempt (observation)',  'SC03 is immediately after SC02 — easy to confuse as "the cloud scene"',
     'Identifies cloud context but cannot pinpoint the exact initiating moment', 'attempt'),
    ('B', 'OG0005_SC02_I', 'Rainbow stardust cloud drifts toward Tiny Rock','YES',100,
     'Initiating Event',       '—',  '—',  'initiating_event'),
    ('C', 'OG0005_SC05_I', 'Didi traps the cloud in a jar',                 'NO',  15,
     'Attempt (capture)',      'Also involves cloud but is the action scene — confuses problem with attempt',
     'Understands cloud is central; cannot map sentence to specific moment', 'attempt'),
    ('D', 'OG0005_SC09_I', 'Podo explains to Didi about light',             'NO',   5,
     'Consequence / Turning Point', 'Completely different scene; student likely guessing',
     'No literal comprehension of the sentence', 'consequence'),
]
for i, (opt, img, desc, correct, score, role, rat, weak, lrs_sg) in enumerate(q2_data):
    r = 6 + i
    sf = score_fill(score)
    sc(ws, r, 1, opt,     fill=sf,   font=FD(9))
    sc(ws, r, 2, img,     fill=YELLOW if correct=='YES' else None, font=FD(9))
    sc(ws, r, 3, desc,    fill=YELLOW if correct=='YES' else None, font=FD(9))
    sc(ws, r, 4, correct, fill=sf,   font=FD(9))
    sc(ws, r, 5, score,   fill=sf,   font=FD(9))
    sc(ws, r, 6, role,    fill=YELLOW if correct=='YES' else None, font=FD(9))
    sc(ws, r, 7, rat,     fill=YELLOW if correct=='YES' else None, font=FD(9))
    sc(ws, r, 8, weak,    fill=YELLOW if correct=='YES' else None, font=FD(9))
    sc(ws, r, 9, lrs_sg,  fill=YELLOW if correct=='YES' else None, font=FD(9))

lrs = ws.cell(11, 1, 'LRS: verb="answered" | object="quiz_OG0005_Q02_sent_match" | result.sg_element="initiating_event" | result.option_selected=<A/B/C/D> | result.score_raw=<n>')
lrs.fill = LT_GREEN; lrs.font = FW(sz=10); lrs.alignment = ALIGN_CW; lrs.border = BORD
ws.merge_cells('A11:I11')

# ══════════════════════════════════════════════════════════
# Q03_UNSCRAMBLE
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('Q03_UNSCRAMBLE')
set_cols(ws, {'A':6,'B':18,'C':18,'D':14,'E':14,'F':14,'G':18,'H':22,'I':24})
set_rows(ws, {1:26,2:22,3:22,4:22,6:22,
              8:30,9:30,10:30,11:30,12:30,13:30,14:30,
              16:22,18:22,19:22,20:22,21:22,22:22,23:22,25:22})

title(ws, 1, 1, 'Q03 — Sentence Unscramble | Story Grammar: Attempt')
ws.merge_cells('A1:I1')
sc(ws, 2, 1, 'Question: The words below are mixed up. Drag them into the correct order to make a sentence.',
   fill=YELLOW, font=Font(bold=True, name='Calibri', size=9))
ws.merge_cells('A2:I2')
sc(ws, 3, 1, 'Source: SC05 narration  |  Correct: "By trapping it, you extinguished its light."',
   fill=YELLOW)
ws.merge_cells('A3:I3')
sc(ws, 4, 1, 'Scrambled tokens presented to student: [ By ] [ trapping ] [ it, ] [ you ] [ extinguished ] [ its ] [ light. ]  (7 tokens — Level 2)',
   fill=LT_BLUE)
ws.merge_cells('A4:I4')

banner(ws, 6, 1, 'SECTION A — Per-Word Position Weight')
ws.merge_cells('A6:I6')
for c, h in enumerate(['Word Token','Correct Pos','Grammatical Role','Word Weight',
                        'Weight Rationale','Max Points','Cumulative'], 1):
    hdr(ws, 7, c, h)

# All word rows → YELLOW; Max Points col → LT_GREEN
word_data = [
    ('By',           1, 'Preposition (adverbial clause opener)',    1.5,
     'Marks adverbial structure; wrong placement = cannot parse complex sentence opener',
     '=D8/SUM($D$8:$D$14)*100',  '=F8'),
    ('trapping',     2, 'Gerund / main verb of subordinate clause', 2.5,
     'High-value academic verb; core meaning word — misplacing = semantic breakdown',
     '=D9/SUM($D$8:$D$14)*100',  '=F8+F9'),
    ('it,',          3, 'Object pronoun of subordinate clause',     1.0,
     'Shortest/lightest token; confusion here is forgivable — function word',
     '=D10/SUM($D$8:$D$14)*100', '=F8+F9+F10'),
    ('you',          4, 'Subject of main clause',                   1.5,
     'Pronoun subject; swapping with "its" = pronoun case confusion',
     '=D11/SUM($D$8:$D$14)*100', '=F8+F9+F10+F11'),
    ('extinguished', 5, 'Main verb (past tense, high-level vocab)', 2.5,
     'Highest-difficulty word; correct placement shows vocabulary + syntax mastery',
     '=D12/SUM($D$8:$D$14)*100', '=F8+F9+F10+F11+F12'),
    ('its',          6, 'Possessive pronoun (modifying "light")',   1.5,
     'Determiner; swapping with "you" = common pronoun confusion',
     '=D13/SUM($D$8:$D$14)*100', '=F8+F9+F10+F11+F12+F13'),
    ('light.',       7, 'Object noun / sentence closer',            2.0,
     'Closing content word; missing = incomplete predicate understanding',
     '=D14/SUM($D$8:$D$14)*100', '=SUM(F8:F14)'),
]
for i, (word, pos, role, wt, rat, formula, cumul) in enumerate(word_data):
    r = 8 + i
    sc(ws, r, 1, word,    fill=YELLOW)
    sc(ws, r, 2, pos,     fill=None,    align=ALIGN_C)
    sc(ws, r, 3, role,    fill=YELLOW)
    sc(ws, r, 4, wt,      fill=None,    align=ALIGN_C)
    sc(ws, r, 5, rat,     fill=None)
    sc(ws, r, 6, formula, fill=LT_GREEN, align=ALIGN_C)
    sc(ws, r, 7, cumul,   fill=None,    align=ALIGN_C)

# TOTAL
sc(ws, 15, 1, 'TOTAL',          fill=LT_BLUE, font=Font(bold=True,name='Calibri',size=9), align=ALIGN_C)
sc(ws, 15, 4, '=SUM(D8:D14)',   fill=LT_BLUE, font=Font(bold=True,name='Calibri',size=9), align=ALIGN_C)
sc(ws, 15, 6, '=SUM(F8:F14)',   fill=LT_BLUE, font=Font(bold=True,name='Calibri',size=9), align=ALIGN_C)
sc(ws, 15, 7, '100 check',      fill=LT_BLUE, font=Font(bold=True,name='Calibri',size=9), align=ALIGN_C)

banner(ws, 16, 1, 'SECTION B — Partial Score Examples')
ws.merge_cells('A16:I16')
for c, h in enumerate(['Student Answer','Correct Positions','Score','Weakness Signal'], 1):
    hdr(ws, 17, c, h, fill=LT_BLUE, fc='000000')

ex_data = [
    ('By trapping it, you extinguished its light.',     '7/7',                       100, '—'),
    ('By trapping it, you extinguished light. its',     '6/7 (its/light swapped)',    84,  'Possessive determiner closing structure'),
    ('trapping By it, you extinguished its light.',     '5/7 (By/trapping swapped)',  68,  'Adverbial clause opener syntax'),
    ('By trapping it, its you extinguished light.',     '4/7 (pronoun confusion)',    52,  'Pronoun + verb-object ordering'),
    ('you extinguished its light. By trapping it,',     '2/7 (clauses reversed)',     28,  'Complex sentence clause ordering'),
    ('light. its extinguished you it, trapping By',     '0/7',                         0,  'No syntactic understanding; likely guessing'),
]
for i, (ans, pos_str, score, weak) in enumerate(ex_data):
    r = 18 + i
    sc(ws, r, 1, ans,     font=FD(9))
    sc(ws, r, 2, pos_str, font=FD(9))
    sc(ws, r, 3, score,   fill=score_fill(score), font=FD(9))
    sc(ws, r, 4, weak,    font=FD(9))

lrs = ws.cell(25, 1, 'LRS: verb="answered" | object="quiz_OG0005_Q03_unscramble" | result.sg_element="attempt" | result.word_order_submitted=[] | result.score_raw=<n> | result.words_correct=<n>/7')
lrs.fill = LT_GREEN; lrs.font = FW(sz=10); lrs.alignment = ALIGN_CW; lrs.border = BORD
ws.merge_cells('A25:I25')

# ══════════════════════════════════════════════════════════
# Q04_EMOTION  (mirror of OG0021 Q06_EMOTION — has source row)
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('Q04_EMOTION')
set_cols(ws, {'A':8,'B':36,'C':12,'D':20,'E':20,'F':24,'G':24,'H':22,'I':20})
set_rows(ws, {1:26,3:22,5:30,6:40,7:40,8:40,9:40,11:22})

title(ws, 1, 1, 'Q04 — Emotion Identification MCQ | Story Grammar: Reaction / Internal Response')
ws.merge_cells('A1:I1')
sc(ws, 2, 1, 'Question: Look at scene SC08. How does Didi feel in this scene?',
   fill=YELLOW, font=Font(bold=True, name='Calibri', size=9))
ws.merge_cells('A2:I2')
sc(ws, 3, 1, 'Source: SC08_Emotion = Disappointed  |  Scene: Didi sees the rainbow cloud has turned completely gray in the jar.',
   fill=LT_BLUE)
ws.merge_cells('A3:I3')

for c, h in enumerate(['Option','Option Text','Correct?','Score','Story Grammar Role',
                        'Distractor Type','Distractor Rationale','Weakness Signal','LRS sg_element'], 1):
    hdr(ws, 5, c, h)

q4_data = [
    ('A', 'Frustrated',    'NO',  50, 'Internal Response (near-miss)',
     'Adjacent negative emotion',
     'Recognizes negative valence but misreads intensity level — frustration implies blocked goal, disappointment implies lost hope',
     'Emotion at valence level only; cannot distinguish frustrated vs disappointed', 'internal_response'),
    ('B', 'Disappointed',  'YES',100, 'Reaction / Internal Response',
     'Correct', 'SC08 emotion confirmed — cloud turned gray, hope lost', '—', 'reaction'),
    ('C', 'Happy',         'NO',   0, 'Internal Response mismatch',
     'Opposite emotion', 'Confuses resolution scene emotion with SC08',
     'Cannot map emotions to specific moments; temporal confusion', 'internal_response'),
    ('D', 'Shocked',       'NO',  20, 'Surface reaction read',
     'Initial reaction distractor',
     'Shock is momentary; disappointed is sustained — student reads surface not depth',
     'Reads immediate facial cue; misses sustained emotional state', 'reaction'),
]
for i, (opt, text, correct, score, role, dtype, rat, weak, lrs_sg) in enumerate(q4_data):
    r = 6 + i
    sf = score_fill(score)
    sc(ws, r, 1, opt,     fill=sf)
    sc(ws, r, 2, text,    font=FD(9))
    sc(ws, r, 3, correct, fill=sf)
    sc(ws, r, 4, score,   fill=sf)
    sc(ws, r, 5, role,    font=FD(9))
    sc(ws, r, 6, dtype,   font=FD(9))
    sc(ws, r, 7, rat,     font=FD(9))
    sc(ws, r, 8, weak,    font=FD(9))
    sc(ws, r, 9, lrs_sg,  font=FD(9))

lrs = ws.cell(11, 1, 'LRS: verb="answered" | object="quiz_OG0005_Q04_emotion" | result.sg_element="reaction" | result.option_selected=<A/B/C/D> | result.score_raw=<n> | result.semantic_distance=<0.0-1.0>')
lrs.fill = LT_GREEN; lrs.font = FW(sz=10); lrs.alignment = ALIGN_CW; lrs.border = BORD
ws.merge_cells('A11:I11')

# ══════════════════════════════════════════════════════════
# Q05_CHAR_GOAL  (mirror of OG0021 Q05_CHAR_GOAL)
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('Q05_CHAR_GOAL')
set_cols(ws, {'A':8,'B':36,'C':12,'D':20,'E':20,'F':24,'G':24,'H':22,'I':20})
set_rows(ws, {1:26,4:30,5:36,6:36,7:36,8:36,10:22})

title(ws, 1, 1, 'Q05 — Character Goal MCQ | Story Grammar: Initiating Event / Goal')
ws.merge_cells('A1:I1')
sc(ws, 2, 1, "What does Didi want at the start of the story?",
   fill=YELLOW, font=Font(bold=True, name='Calibri', size=9))
ws.merge_cells('A2:I2')

for c, h in enumerate(['Option','Option Text','Correct?','Score','Story Grammar Role',
                        'Distractor Type','Distractor Rationale','Weakness Signal','LRS sg_element'], 1):
    hdr(ws, 4, c, h)

q5_data = [
    ('A', 'To capture a piece of the rainbow cloud.',  'YES',100, 'Initiating Event / Goal',
     'Correct', '—', '—', 'initiating_event'),
    ('B', 'To travel across the universe with Podo.', 'NO',   0, 'Unrelated detail',
     'Wild distractor', 'Podo is present but travel is never a stated goal',
     'Random selection or no character goal understanding', '—'),
    ('C', "To share the cloud's beauty with friends.", 'NO',  20, 'Goal misread (altruistic reframe)',
     'Partial goal', "Student softens Didi's selfish possession drive into sharing",
     'Cannot distinguish "keep for self" from "share" — moral reframing of goal', 'initiating_event'),
    ('D', 'To learn why stardust clouds glow.',        'NO',  10, 'Curiosity misread',
     'Cognitive distractor', 'Interprets action as scientific curiosity rather than possessive desire',
     'Reads surface behavior not character motivation; inferencing gap', 'initiating_event'),
]
for i, (opt, text, correct, score, role, dtype, rat, weak, lrs_sg) in enumerate(q5_data):
    r = 5 + i
    sf = score_fill(score)
    sc(ws, r, 1, opt,     fill=sf)
    sc(ws, r, 2, text,    font=FD(9))
    sc(ws, r, 3, correct, fill=sf)
    sc(ws, r, 4, score,   fill=sf)
    sc(ws, r, 5, role,    font=FD(9))
    sc(ws, r, 6, dtype,   font=FD(9))
    sc(ws, r, 7, rat,     font=FD(9))
    sc(ws, r, 8, weak,    font=FD(9))
    sc(ws, r, 9, lrs_sg,  font=FD(9))

lrs = ws.cell(10, 1, 'LRS: verb="answered" | object="quiz_OG0005_Q05_char_goal" | result.sg_element="initiating_event" | result.option_selected=<A/B/C/D> | result.score_raw=<n>')
lrs.fill = LT_GREEN; lrs.font = FW(sz=10); lrs.alignment = ALIGN_CW; lrs.border = BORD
ws.merge_cells('A10:I10')

# ══════════════════════════════════════════════════════════
# Q06_MAIN_IDEA  (mirror of OG0021 Q04_MAIN_IDEA)
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('Q06_MAIN_IDEA')
set_cols(ws, {'A':8,'B':36,'C':12,'D':20,'E':20,'F':24,'G':24,'H':22,'I':20})
set_rows(ws, {1:26,4:30,5:36,6:36,7:36,8:36,10:22})

title(ws, 1, 1, 'Q06 — Main Idea MCQ | Story Grammar: Theme / Resolution')
ws.merge_cells('A1:I1')
sc(ws, 2, 1, 'Question: What lesson did Didi learn from this experience?',
   fill=YELLOW, font=Font(bold=True, name='Calibri', size=9))
ws.merge_cells('A2:I2')

for c, h in enumerate(['Option','Option Text','Correct?','Score','Story Grammar Role',
                        'Distractor Type','Distractor Rationale','Weakness Signal','LRS sg_element'], 1):
    hdr(ws, 4, c, h)

q6_data = [
    ('A', 'Stardust clouds are made of beautiful colors.',             'NO',  10,
     'Setting detail', 'Literal distractor',
     'Takes visual/factual description as message; reading at surface level',
     'Cannot abstract theme from plot; literal reading only', 'setting'),
    ('B', 'True beauty belongs to the open sky, not a shelf.',         'YES',100,
     'Theme / Resolution', 'Correct',
     'Correctly captures freedom-vs-possession arc from SC05 through SC12', '—', 'theme'),
    ('C', "It is important to always listen to your friend's advice.", 'NO',  20,
     'Attempt / Relationship misread', 'Partial theme (process focus)',
     'Podo gives advice which Didi ignores — student focuses on interpersonal lesson not deeper beauty/freedom theme',
     'Mid-level comprehension; identifies supporting lesson not central theme', 'attempt'),
    ('D', 'Capturing rare things is a way to keep their beauty forever.','NO',  0,
     'Anti-theme', 'Opposite meaning',
     "Directly contradicts the story's message — student may not have processed the consequence arc",
     'No thematic understanding; possibly read only first half of story', '—'),
]
for i, (opt, text, correct, score, role, dtype, rat, weak, lrs_sg) in enumerate(q6_data):
    r = 5 + i
    sf = score_fill(score)
    sc(ws, r, 1, opt,     fill=sf)
    sc(ws, r, 2, text,    font=FD(9))
    sc(ws, r, 3, correct, fill=sf)
    sc(ws, r, 4, score,   fill=sf)
    sc(ws, r, 5, role,    font=FD(9))
    sc(ws, r, 6, dtype,   font=FD(9))
    sc(ws, r, 7, rat,     font=FD(9))
    sc(ws, r, 8, weak,    font=FD(9))
    sc(ws, r, 9, lrs_sg,  font=FD(9))

lrs = ws.cell(10, 1, 'LRS: verb="answered" | object="quiz_OG0005_Q06_main_idea" | result.sg_element="theme" | result.option_selected=<A/B/C/D> | result.score_raw=<n> | result.semantic_distance=<0.0-1.0>')
lrs.fill = LT_GREEN; lrs.font = FW(sz=10); lrs.alignment = ALIGN_CW; lrs.border = BORD
ws.merge_cells('A10:I10')

# ══════════════════════════════════════════════════════════
# SCORING_RULES
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('SCORING_RULES')
set_cols(ws, {'A':10,'B':24,'C':30,'D':18,'E':18,'F':28,'G':28})
set_rows(ws, {1:26,3:22,5:28,6:28,7:28,8:28,10:22,
              12:26,13:26,14:26,15:26,16:26,17:26,18:26,
              20:22,21:28,22:28,23:28})

title(ws, 1, 1, 'Global Scoring Rules — OG0005 Reading Quiz')
ws.merge_cells('A1:G1')

banner(ws, 3, 1, 'SECTION A — Score Band Interpretation')
ws.merge_cells('A3:G3')
for c, h in enumerate(['Band','Score Range','Color','LRS Risk Signal',
                        'SG Insight','MRI Profile Impact','Recommendation'], 1):
    hdr(ws, 4, c, h, fill=LT_BLUE, fc='000000')

bands = [
    ('Mastery',    GREEN,    '85-100', 'Green',  'None',                        'Full SG element understood',        'V+/L+ in profile',            'Proceed to next level'),
    ('Developing', YELLOW,   '60-84',  'Gold',   'Meaning Signal (mild)',        'Partial narrative understanding',   'V0 — maintain level',         'Review target scene'),
    ('Emerging',   PINK_RED, '30-59',  'Orange', 'Meaning Signal (moderate)',    'Surface/literal only',              'V- — flag for booster',       'Assign Booster Content'),
    ('Beginning',  LT_RED,   '0-29',   'Red',    'Meaning Signal (strong)',      'Fundamental gap; guessing',         'V-- — priority intervention', 'Comprehension Booster + re-read'),
]
for i, (band, bfill, rng, color, lrs_sig, sg_insight, mri, rec) in enumerate(bands):
    r = 5 + i
    sc(ws, r, 1, band,     font=FD(9))
    sc(ws, r, 2, rng,      fill=bfill, font=FD(9))
    sc(ws, r, 3, color,    font=FD(9))
    sc(ws, r, 4, lrs_sig,  font=FD(9))
    sc(ws, r, 5, sg_insight, font=FD(9))
    sc(ws, r, 6, mri,      font=FD(9))
    sc(ws, r, 7, rec,      font=FD(9))

banner(ws, 10, 1, 'SECTION B — Weighting Rationale by Story Grammar Element')
ws.merge_cells('A10:G10')
for c, h in enumerate(['SG Element','Quiz Coverage','Why Weighted','LRS Engine','Weakness Signal if Missed'], 1):
    hdr(ws, 11, c, h, fill=LT_BLUE, fc='000000')

sg_rows = [
    ('Setting / Initiating Event (SC02)', 'Q01 (SC02, anchor pos 1)', 'Entry anchor — gate to story world and all events; SC02 IS the initiating event', 'Vocab & Comprehension Engine', 'Literal comprehension dependency'),
    ('Initiating Event / Goal',           'Q01 (SC03), Q02, Q05',     'Core problem-formation; misidentifying = no narrative schema',                    'Vocab & Comprehension Engine', 'Inferencing weakness'),
    ('Attempt (cloud capture)',           'Q01 (SC05), Q03',          'Complex sentence decoding (By trapping it...); confusion = sequential gap + vocab', 'Vocab & Comprehension Engine', 'Semantic Decision Stability + Vocab gap'),
    ('Consequence / Reaction',            'Q01 (SC07), Q04',          'Emotional scene + cause-effect link — confusion = emotion inference gap',           'Expression Engine',             'Emotion inference (Expression Signal)'),
    ('Resolution (cloud freed)',          'Q01 (SC12, anchor pos 5), Q06', 'Exit anchor — without this, theme extraction fails',                           'Vocab & Comprehension Engine', 'Inferencing / Story Transfer weakness'),
    ('Theme',                             'Q06',                      'Highest-order comprehension — requires abstract reasoning beyond literal events',   'V&C Engine + Expression Engine','Limited reasoning (Expression Signal)'),
    ('Internal Response (Emotion)',       'Q04',                      'Moment-level emotion mapping at advanced vocabulary level',                         'Expression Engine',             'Perspective taking + emotion vocabulary weakness'),
]
for i, row in enumerate(sg_rows):
    r = 12 + i
    sc(ws, r, 1, row[0], fill=YELLOW, font=FD(9))
    for c, v in enumerate(row[1:], 2):
        sc(ws, r, c, v, font=FD(9))

banner(ws, 20, 1, 'SECTION C — MRI Profile Contribution')
ws.merge_cells('A20:G20')
sc(ws, 21, 1, 'Quiz scores feed into Vocabulary & Comprehension Engine (V-score) and partially Expression Engine (L-score).', fill=LT_BLUE, font=FD(9))
ws.merge_cells('A21:G21')
sc(ws, 22, 1, 'Parent Report (3-tier): Level (e.g. Lv4 — Vocabulary Retrieval) | Detail Profile: P3/V1/L2 | Risk Points: Inferencing weakness / Emotion identification gap', fill=LT_GREEN, font=FD(9))
ws.merge_cells('A22:G22')
sc(ws, 23, 1, "OG0005 Q03 uses a 7-word adverbial-clause sentence vs OG0021's 5-word simple sentence (Level 2). Total weight = 12.5. Academic verbs 'trapping' and 'extinguished' carry highest weights (2.5 each).", fill=LT_BLUE, font=FD(9))
ws.merge_cells('A23:G23')

# ══════════════════════════════════════════════════════════
# LRS_MAPPING
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet('LRS_MAPPING')
set_cols(ws, {'A':8,'B':20,'C':28,'D':28,'E':16,'F':16,'G':28,'H':30})
set_rows(ws, {1:26,2:28,3:28,4:28,5:28,6:28,7:28,8:28,10:30})

title(ws, 1, 1, 'LRS xAPI Mapping — OG0005 Quiz Statements')
ws.merge_cells('A1:H1')
for c, h in enumerate(['Q_ID','xAPI Verb','xAPI Object','result.sg_element',
                        'result.score_raw','result.correct','Extra Fields','Risk Signal'], 1):
    hdr(ws, 2, c, h)

lrs_rows = [
    ('Q01','answered','quiz_OG0005_Q01_sequencing',
     'setting,initiating_event,attempt,consequence,resolution',
     '0-100','partial','sequence_submitted[], position_scores{}','Meaning Signal if <60'),
    ('Q02','answered','quiz_OG0005_Q02_sent_match',
     'initiating_event','0-100','true/false',
     'option_selected, response_latency_ms','Meaning Signal if option!=B'),
    ('Q03','answered','quiz_OG0005_Q03_unscramble',
     'attempt','0-100','partial',
     'word_order_submitted[], words_correct/7','Meaning Signal if <60'),
    ('Q04','answered','quiz_OG0005_Q04_emotion',
     'reaction','0-100','true/false',
     'option_selected, semantic_distance','Expression Signal if <40 (emotion inference)'),
    ('Q05','answered','quiz_OG0005_Q05_char_goal',
     'initiating_event','0-100','true/false',
     'option_selected, response_latency_ms','Meaning Signal if option!=A'),
    ('Q06','answered','quiz_OG0005_Q06_main_idea',
     'theme','0-100','true/false',
     'option_selected, semantic_distance','Expression Signal if <30 (limited reasoning)'),
]
for i, row in enumerate(lrs_rows):
    r = 3 + i
    sc(ws, r, 1, row[0], fill=YELLOW, font=FD(9))
    for c, v in enumerate(row[1:], 2):
        sc(ws, r, c, v, font=FD(9))

note = ws.cell(10, 1, 'Single LRS storage (xAPI). V&C Engine reads sg_element + score_raw for Q01-Q05. Expression Engine reads Q04/Q06 semantic_distance. MRI report: V-level (Q01-Q05) + L-level (Q04, Q06).')
note.fill = LT_GREEN; note.font = FD(9); note.alignment = ALIGN_LW; note.border = BORD
ws.merge_cells('A10:H10')

# ══════════════════════════════════════════════════════════
out = r'C:\Users\bonni\Desktop\ISM\Content\Quiz\OG0005\OG0005_ReadingQuiz.xlsx'
wb.save(out)
print('Saved:', out)
