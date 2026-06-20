from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "OG0021_ReadingQuiz.xlsx"

C_HEADER = "1F4E79"
C_SUB = "2E75B6"
C_ACCENT = "BDD7EE"
C_YELLOW = "FFF2CC"
C_GREEN = "E2EFDA"
C_ORANGE = "FCE4D6"
C_OK = "70AD47"
C_PART = "FFD966"
C_LOW = "FFEECC"
C_BAD = "FFCCCC"


def style_cell(c, fill=None, bold=False, color="000000", size=9, align="left"):
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    return c


def header(ws, row, col, value, fill=C_SUB, size=10):
    return style_cell(ws.cell(row, col, value), fill=fill, bold=True, color="FFFFFF", size=size, align="center")


def merge_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    style_cell(c, C_HEADER, True, "FFFFFF", 12, "center")
    c.value = text


def set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def option_fill(score):
    if score == 100:
        return C_OK
    if score >= 30:
        return C_PART
    if score > 0:
        return C_LOW
    return C_BAD


wb = Workbook()

# QUIZ_LIST
ws = wb.active
ws.title = "QUIZ_LIST"
set_widths(ws, [8, 20, 22, 30, 18, 18, 12, 22, 22, 18])
merge_title(ws, "A1:J1", "OG0021 Reading Quiz v2 — Story Grammar + Synthesis")
ws.row_dimensions[1].height = 28
cols = ["Q_ID", "Quiz Type", "Primary Story Grammar", "Question (EN)", "Resource", "Correct Answer", "Max Score", "LRS sg_element", "Scoring Mode", "Sheet Ref"]
for i, h in enumerate(cols, 1):
    header(ws, 2, i, h)
ws.row_dimensions[2].height = 30
rows = [
    ["Q01", "Setting Image Choice", "Setting", "Where does Milo go to look for his color?", "SC01/03/06/10 images", "Option B", 100, "setting", "Weighted MCQ", "Q01_SETTING"],
    ["Q02", "Listening Scene Match", "Initiating Event", "Listen. Which scene starts the problem?", "Audio + SC02/03/06/09", "Option A", 100, "initiating_event", "Weighted MCQ", "Q02_INIT_EVENT"],
    ["Q03", "Scene-Anchored Unscramble", "Attempt", "Put the story words in order.", "SC03 image", "Milo walks into the forest.", 100, "attempt", "Weighted Unscramble", "Q03_ATTEMPT"],
    ["Q04", "Feeling Match", "Reaction", "How does Milo feel here?", "SC06 image", "Option B", 100, "reaction", "Weighted MCQ", "Q04_REACTION"],
    ["Q05", "Internal Response MCQ", "Internal Response", "What is Milo thinking?", "SC06 image", "Option A", 100, "internal_response", "Weighted MCQ", "Q05_INTERNAL"],
    ["Q06", "Ending Scene Sequence", "Consequence", "Put the ending scenes in order.", "SC06/07/08/09 images", "SC06 → SC07 → SC08 → SC09", 100, "consequence", "Weighted Position", "Q06_CONSEQUENCE"],
    ["Q07", "Synthesis MCQ", "Synthesis", "What did Milo find out at the end?", "-", "Option C", 100, "synthesis", "Weighted MCQ", "Q07_SYNTHESIS"],
]
for r, row in enumerate(rows, 3):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v))
    ws.row_dimensions[r].height = 28
ws.merge_cells("A11:J11")
style_cell(ws["A11"], C_GREEN, False, "000000", 9, "left").value = "v2 keeps one primary Story Grammar element per scored question. Q7 Synthesis is reported separately from the six-axis graph."


def make_option_sheet(title, sheet, question, options, lrs, resource="-"):
    ws = wb.create_sheet(sheet)
    set_widths(ws, [8, 36, 12, 20, 20, 24, 28, 28, 22])
    merge_title(ws, "A1:I1", title)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:I2")
    style_cell(ws["A2"], C_YELLOW, True, "000000", 9).value = "Question: " + question
    if resource != "-":
        ws.merge_cells("A3:I3")
        style_cell(ws["A3"], C_ACCENT, False, "000000", 9).value = "Resource: " + resource
    hdrs = ["Option", "Option Text", "Is Correct?", "Score", "SG Element", "Distractor Type", "Distractor Rationale", "Student Weakness Signal", "LRS sg_element"]
    start = 5 if resource != "-" else 4
    for i, h in enumerate(hdrs, 1):
        header(ws, start, i, h)
    for r, row in enumerate(options, start + 1):
        fill = option_fill(row[3])
        for c, v in enumerate(row, 1):
            bg = fill if c in (1, 3, 4) else None
            style_cell(ws.cell(r, c, v), bg, c in (1, 3, 4) and row[3] == 100, "FFFFFF" if bg == C_OK and c in (1, 3, 4) else "000000")
        ws.row_dimensions[r].height = 36
    foot = start + len(options) + 2
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=9)
    style_cell(ws.cell(foot, 1), C_GREEN, True, "000000", 9).value = f'LRS xAPI: verb="answered" | object="quiz_OG0021_{sheet.lower()}" | result.sg_element="{lrs}" | result.score_raw=<pts>'


# Q01 Setting
ws = wb.create_sheet("Q01_SETTING")
set_widths(ws, [8, 30, 12, 18, 18, 28, 30, 24, 22])
merge_title(ws, "A1:I1", "Q01 — Setting Image Choice | Story Grammar: Setting")
ws.row_dimensions[1].height = 26
ws.merge_cells("A2:I2")
style_cell(ws["A2"], C_YELLOW, True).value = "Question: Where does Milo go to look for his color?"
for i, h in enumerate(["Option", "Scene / Resource", "Correct?", "Score", "SG Role", "Distractor Rationale", "Student Weakness Signal", "Error Tag", "LRS"], 1):
    header(ws, 4, i, h)
data = [
    ["A", "OG0021_SC01_I", "NO", 35, "Opening character", "인물 소개 장면과 실제 탐색 장소를 혼동함", "Opening/setting detail confusion", "setting_actor_scene", "setting"],
    ["B", "OG0021_SC03_I", "YES", 100, "Place setting", "Milo walks into the forest.", "-", "correct_setting_place", "setting"],
    ["C", "OG0021_SC06_I", "NO", 55, "Later place", "중요 장소인 연못을 탐색이 시작된 장소와 혼동함", "Later consequence place confusion", "setting_later_place", "setting"],
    ["D", "OG0021_SC10_I", "NO", 20, "Ending place", "결말의 집 장면을 주요 탐색 장소로 혼동함", "Ending/opening place confusion", "setting_ending_place", "setting"],
]
for r, row in enumerate(data, 5):
    fill = option_fill(row[3])
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), fill if c in (1, 3, 4) else None, c in (1, 3, 4) and row[3] == 100, "FFFFFF" if fill == C_OK and c in (1, 3, 4) else "000000")
    ws.row_dimensions[r].height = 30
ws.merge_cells("A11:I11")
style_cell(ws["A11"], C_GREEN, True).value = 'LRS xAPI: verb="answered" | object="quiz_OG0021_v2_q01_setting" | result.sg_element="setting" | result.score_raw=<pts>'

# Q02
make_option_sheet(
    "Q02 — Listening Scene Match | Story Grammar: Initiating Event",
    "Q02_INIT_EVENT",
    "Listen. Which scene starts the problem?",
    [
        ["A", "OG0021_SC02_I", "YES", 100, "Initiating Event", "Correct", "Milo wakes up gray; the story problem begins.", "-", "initiating_event"],
        ["B", "OG0021_SC03_I", "NO", 20, "Attempt", "Later action", "Student chooses the search after the problem, not the problem itself.", "Event/action sequence confusion", "attempt"],
        ["C", "OG0021_SC06_I", "NO", 30, "Reaction", "Result scene", "Student focuses on sadness after the problem.", "Initiating event vs reaction confusion", "reaction"],
        ["D", "OG0021_SC09_I", "NO", 0, "Resolution", "Opposite phase", "Student chooses a late recovery scene.", "Story direction weakness", "consequence"],
    ],
    "initiating_event",
    "Audio/OG0021_SC02_ST01_N_A.mp3",
)

# Q03
ws = wb.create_sheet("Q03_ATTEMPT")
set_widths(ws, [6, 18, 18, 14, 14, 16, 18, 24, 24])
merge_title(ws, "A1:I1", "Q03 — Scene-Anchored Unscramble | Story Grammar: Attempt")
ws.merge_cells("A2:I2")
style_cell(ws["A2"], C_YELLOW, True).value = "Question: Put the story words in order."
ws.merge_cells("A3:I3")
style_cell(ws["A3"], C_YELLOW).value = 'Correct sentence from SC03_ST01_N: "Milo walks into the forest."'
ws.merge_cells("A4:I4")
style_cell(ws["A4"], C_ACCENT).value = "Scrambled words: walks | forest. | Milo | the | into"
for i, h in enumerate(["Word", "Correct Pos", "Grammar Role", "Word Weight", "Weight Rationale", "Max Points", "Error Tag"], 1):
    header(ws, 6, i, h)
words = [
    ["Milo", 1, "Subject", 1.5, "Identifies actor of attempt.", "=D7/SUM($D$7:$D$12)*100", "actor_order"],
    ["walks", 2, "Action verb", 2.5, "Core attempt action; highest weight.", "=D8/SUM($D$7:$D$11)*100", "attempt_action"],
    ["into", 3, "Direction word", 1.5, "Connects action to place.", "=D9/SUM($D$7:$D$11)*100", "direction_word"],
    ["the", 4, "Article", 1.0, "Small syntax support.", "=D10/SUM($D$7:$D$11)*100", "article"],
    ["forest.", 5, "Place noun", 2.5, "Shows where the attempt begins.", "=D11/SUM($D$7:$D$11)*100", "attempt_place"],
]
for r, row in enumerate(words, 7):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), C_GREEN if c == 6 else (C_YELLOW if c in (1, 3) else None), False, "000000", 9, "center" if c in (2, 4, 6) else "left")
    ws.row_dimensions[r].height = 30
for c, v in [(1, "TOTAL"), (4, "=SUM(D7:D11)"), (6, "=SUM(F7:F11)")]:
    style_cell(ws.cell(12, c, v), C_ACCENT, True, "000000", 9, "center")
ws.merge_cells("A15:I15")
header(ws, 15, 1, "SECTION B — Partial Score Examples")
for i, h in enumerate(["Student Answer", "Score", "Interpretation", "Weakness Signal"], 1):
    header(ws, 16, i, h, C_ACCENT)
examples = [
    ["Milo walks into the forest.", 100, "Exact story sentence restored.", "-"],
    ["Milo walks the into forest.", 83, "Core attempt understood; small syntax issue.", "function word placement"],
    ["forest. into the walks Milo", 17, "Place is known but action sequence is broken.", "attempt sequence weakness"],
]
for r, row in enumerate(examples, 17):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), option_fill(v) if c == 2 else None)

# Q04-Q07
make_option_sheet("Q04 — Feeling Match | Story Grammar: Reaction", "Q04_REACTION", "How does Milo feel here?", [
    ["A", "Happy", "NO", 0, "Reaction", "Opposite emotion", "Confuses resolution emotion with sadness.", "Temporal emotion confusion", "reaction"],
    ["B", "Sad", "YES", 100, "Reaction", "Correct", "SC06 shows crying and sadness.", "-", "reaction"],
    ["C", "Angry", "NO", 40, "Reaction", "Adjacent negative emotion", "Recognizes negative feeling but labels it too strongly.", "Emotion differentiation", "reaction"],
    ["D", "Surprised", "NO", 20, "Reaction", "Early-event emotion", "Confuses surprise with later sadness.", "Scene emotion mapping", "reaction"],
], "reaction", "OG0021_SC06_I.png")

make_option_sheet("Q05 — Internal Response MCQ | Story Grammar: Internal Response", "Q05_INTERNAL", "What is Milo thinking?", [
    ["A", "Everyone has their own color.", "YES", 100, "Internal Response", "Correct", "Uses the actual SC06 thought/sentence to infer Milo's inner state.", "-", "internal_response"],
    ["B", "I want to fly with the butterfly.", "NO", 20, "Attempt detail", "Confuses earlier butterfly scene with Milo's thought.", "Thought/action confusion", "attempt"],
    ["C", "The pond is very blue.", "NO", 45, "Surface observation", "Sees visual detail but not internal state.", "Surface-level inference", "internal_response"],
    ["D", "I do not need my color.", "NO", 0, "Opposite motive", "Contradicts Milo's motivation to find his color.", "Motive inversion", "internal_response"],
], "internal_response", "OG0021_SC06_I.png")

ws = wb.create_sheet("Q06_CONSEQUENCE")
set_widths(ws, [8, 20, 14, 14, 24, 32, 20, 24])
merge_title(ws, "A1:H1", "Q06 — Ending Scene Sequence | Story Grammar: Consequence")
ws.merge_cells("A2:H2")
style_cell(ws["A2"], C_YELLOW, True).value = "Question: Put the ending scenes in order."
for i, h in enumerate(["Scene", "Resource", "Correct Pos", "Scene Weight", "SG Role", "Weight Rationale", "Error Tag", "LRS"], 1):
    header(ws, 4, i, h)
rows = [
    ["SC06", "OG0021_SC06_I", 1, 2.0, "Cause / low point", "Milo cries; this begins the ending chain.", "missed_cause_scene", "consequence"],
    ["SC07", "OG0021_SC07_I", 2, 1.5, "Immediate result", "The pond shines after the tear falls.", "immediate_result_shift", "consequence"],
    ["SC08", "OG0021_SC08_I", 3, 1.5, "Internal turn", "Milo realizes his color may be inside him.", "realization_order", "consequence"],
    ["SC09", "OG0021_SC09_I", 4, 2.5, "Final consequence", "Milo's colors come back; final outcome receives high weight.", "final_result_order", "consequence"],
]
for r, row in enumerate(rows, 5):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), C_YELLOW if row[3] >= 2.0 else C_GREEN)
for c, v in [(1, "TOTAL"), (4, "=SUM(D5:D8)")]:
    style_cell(ws.cell(9, c, v), C_ACCENT, True, "000000", 9, "center")

make_option_sheet("Q07 — Synthesis MCQ | Main Idea / Whole Story Meaning", "Q07_SYNTHESIS", "What did Milo find out at the end?", [
    ["A", "Colors keep you safe.", "NO", 10, "Synthesis", "Fact distractor", "Treats an animal fact as the story meaning.", "Theme/detail confusion", "synthesis"],
    ["B", "Friends always help you.", "NO", 30, "Synthesis", "Partial action theme", "Over-focuses on help/action rather than self-discovery.", "Partial theme", "synthesis"],
    ["C", "Your color is inside you.", "YES", 100, "Synthesis", "Correct", "Synthesizes the story's whole meaning.", "-", "synthesis"],
    ["D", "The world has many colors.", "NO", 20, "Synthesis", "Setting detail", "Chooses atmosphere/background over theme.", "Detail/theme confusion", "synthesis"],
], "synthesis")

# SG_SCORING
ws = wb.create_sheet("SG_SCORING")
set_widths(ws, [18, 14, 18, 28, 34, 30, 28])
merge_title(ws, "A1:G1", "Story Grammar Scoring Model — v2")
for i, h in enumerate(["Axis", "Q_ID", "Score Source", "Calculation", "Interpretation", "Parent Report Use", "Weekly Rollup"], 1):
    header(ws, 3, i, h)
axes = [
    ["Setting", "Q01", "setting_score", "weighted option score", "Understands the key place/background for the story action.", "Radar axis 1", "weighted average by story level"],
    ["Initiating Event", "Q02", "initiating_event_score", "option score", "Understands what started the problem.", "Radar axis 2", "weighted average by story level"],
    ["Attempt", "Q03", "attempt_score", "weighted word sequence", "Understands what the character does to solve the problem.", "Radar axis 3", "weighted average by story level"],
    ["Reaction", "Q04", "reaction_score", "option score", "Understands visible feeling/response.", "Radar axis 4", "weighted average by story level"],
    ["Internal Response", "Q05", "internal_response_score", "option score", "Infers thoughts, motive, inner state.", "Radar axis 5", "weighted average by story level"],
    ["Consequence", "Q06", "consequence_score", "weighted scene position", "Understands result and event development.", "Radar axis 6", "weighted average by story level"],
    ["Synthesis", "Q07", "synthesis_score", "option score", "Synthesizes whole-story meaning.", "Separate card; not in radar", "20% of overall reading score"],
]
for r, row in enumerate(axes, 4):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), C_YELLOW if c == 1 else None)
ws.merge_cells("A13:G13")
style_cell(ws["A13"], C_GREEN, True).value = "Overall Reading Score = average(6 Story Grammar axes) × 0.8 + Synthesis × 0.2"

# LRS_MAPPING
ws = wb.create_sheet("LRS_MAPPING")
set_widths(ws, [8, 20, 30, 24, 16, 16, 30, 34])
merge_title(ws, "A1:H1", "LRS xAPI Mapping — OG0021 v2")
for i, h in enumerate(["Q_ID", "xAPI Verb", "xAPI Object", "result.sg_element", "score_raw", "correct", "Extra Fields", "Risk Signal"], 1):
    header(ws, 2, i, h)
for r, row in enumerate([
    ["Q01", "answered", "quiz_OG0021_v2_Q01_setting", "setting", "0-100", "partial", "slot_values, component_scores", "Setting gap if <70"],
    ["Q02", "answered", "quiz_OG0021_v2_Q02_initiating_event", "initiating_event", "0-100", "true/false", "option_selected, audio_src", "Initiating event gap if <70"],
    ["Q03", "answered", "quiz_OG0021_v2_Q03_attempt", "attempt", "0-100", "partial", "word_order, word_scores", "Attempt/action gap if <70"],
    ["Q04", "answered", "quiz_OG0021_v2_Q04_reaction", "reaction", "0-100", "true/false", "option_selected", "Reaction/emotion gap if <70"],
    ["Q05", "answered", "quiz_OG0021_v2_Q05_internal_response", "internal_response", "0-100", "true/false", "thought_selected", "Inference gap if <70"],
    ["Q06", "answered", "quiz_OG0021_v2_Q06_consequence", "consequence", "0-100", "partial", "cause_card,result_card", "Consequence gap if <70"],
    ["Q07", "answered", "quiz_OG0021_v2_Q07_synthesis", "synthesis", "0-100", "true/false", "option_selected", "Synthesis gap if <70"],
], 3):
    for c, v in enumerate(row, 1):
        style_cell(ws.cell(r, c, v), C_YELLOW if c == 1 else None)

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

wb.save(OUT)
print(f"Saved {OUT}")
