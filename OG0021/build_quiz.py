import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

C_HEADER  = '1F4E79'
C_SUB     = '2E75B6'
C_ACCENT  = 'BDD7EE'
C_YELLOW  = 'FFF2CC'
C_GREEN   = 'E2EFDA'
C_ORANGE  = 'FCE4D6'
C_CORRECT = '70AD47'
C_PARTIAL = 'FFD966'

def hdr(ws, row, col, text, bg=C_HEADER, fg='FFFFFF', bold=True, wrap=True, size=10):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    return c

def cell(ws, row, col, text, bg=None, bold=False, wrap=True, align='left', size=9):
    c = ws.cell(row=row, column=col, value=text)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    c.font = Font(bold=bold, size=size)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c

def brd(ws, r1, r2, c1, c2):
    t = Side(style='thin')
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).border = Border(left=t, right=t, top=t, bottom=t)

# ─── SHEET 1: QUIZ_LIST ───────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'QUIZ_LIST'
ws1.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJ', [8,18,22,20,14,14,16,22,22,16]):
    ws1.column_dimensions[col].width = w

ws1.merge_cells('A1:J1')
hdr(ws1, 1, 1, 'OG0021 Reading Quiz — Master List (Milo and the Lost Color)', C_HEADER, size=12)
ws1.row_dimensions[1].height = 28

headers = ['Q_ID','Quiz Type','Story Grammar Element','Question (EN)','Scene Ref','Answer Sheet','Max Score','LRS sg_element','Scoring Mode','Sheet Ref']
for i, h in enumerate(headers, 1):
    hdr(ws1, 2, i, h, C_SUB)
ws1.row_dimensions[2].height = 30

rows = [
    ['Q01','Scene Sequencing','Setting→IE→Attempt→Resolution','Put the 5 scenes in story order.','SC01/SC03/SC05/SC06/SC10','1-3-5-6-10','100','setting,initiating_event,attempt,resolution','Weighted Position','Q01_SEQUENCING'],
    ['Q02','Sentence-Scene Match','Initiating Event','Which scene does this sentence come from?','SC02/SC03/SC06/SC08','SC02','100','initiating_event','Weighted MCQ','Q02_SENT_MATCH'],
    ['Q03','Sentence Unscramble','Attempt','Put the words in order to make a sentence.','SC04_ST02_N','My color helps me fly.','100','attempt','Weighted Unscramble','Q03_UNSCRAMBLE'],
    ['Q04','Main Idea MCQ','Theme / Resolution','What is the main message of this story?','-','Option C','100','theme','Weighted MCQ','Q04_MAIN_IDEA'],
    ['Q05','Character Goal MCQ','Initiating Event / Goal','What does Milo want to find?','-','Option A','100','initiating_event','Weighted MCQ','Q05_CHAR_GOAL'],
    ['Q06','Emotion Identification','Reaction / Internal Response','How does Milo feel in scene SC06?','SC06','Option B','100','reaction','Weighted MCQ','Q06_EMOTION'],
]
for i, r in enumerate(rows, 3):
    for j, v in enumerate(r, 1):
        cell(ws1, i, j, v)
    ws1.row_dimensions[i].height = 28

brd(ws1, 1, 8, 1, 10)

ws1.merge_cells('A10:J10')
n = ws1.cell(10, 1, value='Each question sheet defines per-option/position scores and weights. The HTML file reads this workbook via SheetJS to render the live quiz.')
n.font = Font(italic=True, size=9, color='555555')
n.alignment = Alignment(wrap_text=True, vertical='center')
ws1.row_dimensions[10].height = 20

# ─── SHEET 2: Q01_SEQUENCING ─────────────────────────────────────────────
ws2 = wb.create_sheet('Q01_SEQUENCING')
ws2.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJ', [8,14,30,12,12,14,12,12,20,28]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells('A1:J1')
hdr(ws2, 1, 1, 'Q01 — Scene Sequencing | Story Grammar: Full Arc', C_HEADER, size=11)
ws2.row_dimensions[1].height = 26

ws2.merge_cells('A2:J2')
cell(ws2, 2, 1, 'Question: Put the 5 scene images in the correct story order by dragging them into position.', C_YELLOW, bold=True)
ws2.merge_cells('A3:J3')
cell(ws2, 3, 1, 'Scenes provided:  OG0021_SC01_I  |  OG0021_SC03_I  |  OG0021_SC05_I  |  OG0021_SC06_I  |  OG0021_SC10_I', C_YELLOW)
ws2.merge_cells('A4:J4')
cell(ws2, 4, 1, 'Correct Answer: SC01 → SC03 → SC05 → SC06 → SC10', C_GREEN, bold=True)
ws2.merge_cells('A5:J5')
cell(ws2, 5, 1, 'Story Grammar: SC01=Setting | SC03=Initiating Event/1st Attempt | SC05=2nd Attempt | SC06=Consequence/Reaction | SC10=Resolution', C_ACCENT)
for r in [2,3,4,5]:
    ws2.row_dimensions[r].height = 22

ws2.merge_cells('A7:J7')
hdr(ws2, 7, 1, 'SECTION A — Position Weight per Scene', C_SUB, size=10)
ws2.row_dimensions[7].height = 22

pw_headers = ['Scene Asset','Correct Pos','Story Grammar Role','Position Weight','Weight Rationale','Max Points']
for i, h in enumerate(pw_headers, 1):
    hdr(ws2, 8, i, h, C_SUB)

pw_data = [
    ['OG0021_SC01_I', 1, 'Setting (story opening)', 2.5, 'Unambiguous first scene. Placing wrong = fundamental arc gap.', '=D9/SUM($D$9:$D$13)*100'],
    ['OG0021_SC03_I', 2, 'Initiating Event / 1st Attempt', 1.5, 'Clear event but mid-story; SC03/SC05 confusion is forgivable.', '=D10/SUM($D$9:$D$13)*100'],
    ['OG0021_SC05_I', 3, '2nd Attempt', 1.5, 'Similar to SC03; mixing these two is penalized less.', '=D11/SUM($D$9:$D$13)*100'],
    ['OG0021_SC06_I', 4, 'Consequence / Reaction', 1.0, 'Emotional turning-point; close to SC05 thematically — smallest penalty.', '=D12/SUM($D$9:$D$13)*100'],
    ['OG0021_SC10_I', 5, 'Resolution (returns home)', 2.5, 'Unambiguous final scene. Placing wrong = fundamental arc gap.', '=D13/SUM($D$9:$D$13)*100'],
]
for i, r in enumerate(pw_data, 9):
    bg = C_YELLOW if i in [9, 13] else C_GREEN
    for j, v in enumerate(r, 1):
        c = ws2.cell(i, j, value=v)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal='center' if j in [2,4,6] else 'left', vertical='center', wrap_text=True)
    ws2.row_dimensions[i].height = 30

for col, val in [(1,'TOTAL'), (4,'=SUM(D9:D13)'), (6,'=SUM(F9:F13)')]:
    c = ws2.cell(14, col, value=val)
    c.fill = PatternFill('solid', fgColor=C_ACCENT)
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
brd(ws2, 7, 14, 1, 6)

ws2.merge_cells('A16:J16')
hdr(ws2, 16, 1, 'SECTION B — Score Matrix: Points per (Scene, Submitted Position)', C_SUB, size=10)
ws2.row_dimensions[16].height = 22

ws2.cell(17, 1).value = 'Scene \\ Submitted Pos'
ws2.cell(17, 1).font = Font(bold=True, size=9)
ws2.cell(17, 1).alignment = Alignment(wrap_text=True, vertical='center')
for p in range(1, 6):
    hdr(ws2, 17, p+1, f'Pos {p}', C_SUB)

scenes = ['SC01','SC03','SC05','SC06','SC10']
weights = [2.5, 1.5, 1.5, 1.0, 2.5]
total_w = sum(weights)
correct_pos = [1, 2, 3, 4, 5]

for si, (sc, w, cp) in enumerate(zip(scenes, weights, correct_pos)):
    row = 18 + si
    c = ws2.cell(row, 1, value=f'OG0021_{sc}_I')
    c.fill = PatternFill('solid', fgColor=C_ACCENT)
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='left', vertical='center')
    max_pts = round(w / total_w * 100, 1)
    for p in range(1, 6):
        distance = abs(p - cp)
        pts = round(max_pts * max(0, 1 - distance * 0.5), 1) if distance > 0 else max_pts
        bg = C_CORRECT if p == cp else (C_PARTIAL if pts > 0 else 'FFCCCC')
        c2 = ws2.cell(row, p+1, value=pts if pts > 0 else 0)
        c2.fill = PatternFill('solid', fgColor=bg)
        c2.font = Font(bold=(p==cp), size=9, color='FFFFFF' if p==cp else '000000')
        c2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[row].height = 22

brd(ws2, 17, 22, 1, 6)

ws2.merge_cells('A24:J24')
hdr(ws2, 24, 1, 'SECTION C — Example Answer Scores', C_SUB, size=10)
ws2.row_dimensions[24].height = 22

for i, h in enumerate(['Example Student Answer', 'Submitted Order', 'Score', 'Interpretation'], 1):
    hdr(ws2, 25, i, h, C_ACCENT, '000000')

examples = [
    ['Perfect', 'SC01-SC03-SC05-SC06-SC10', 100, 'Full story arc understood'],
    ['SC05/SC06 swapped', 'SC01-SC03-SC06-SC05-SC10', 80, 'Adjacent attempt/consequence swap — penalty mild'],
    ['Middle scrambled', 'SC01-SC06-SC05-SC03-SC10', 55, 'Anchor points correct; internal sequence weak'],
    ['Beginning/end wrong', 'SC03-SC01-SC10-SC05-SC06', 30, 'Story arc fundamentally misunderstood'],
]
for i, r in enumerate(examples, 26):
    score_bg = C_CORRECT if r[2]==100 else (C_PARTIAL if r[2]>=60 else 'FFCCCC')
    for j, v in enumerate(r, 1):
        bg = score_bg if j == 3 else None
        cell(ws2, i, j, v, bg)
    ws2.row_dimensions[i].height = 20

brd(ws2, 24, 29, 1, 4)

ws2.merge_cells('A31:J31')
hdr(ws2, 31, 1, 'LRS: verb="answered" | object="quiz_OG0021_Q01_sequencing" | result.sg_element="setting,initiating_event,attempt,consequence,resolution" | result.sequence_submitted=[] | result.score_raw=<n> | result.position_scores={}', C_GREEN)
ws2.row_dimensions[31].height = 24
brd(ws2, 31, 31, 1, 10)

# ─── SHEET 3: Q02_SENT_MATCH ─────────────────────────────────────────────
ws3 = wb.create_sheet('Q02_SENT_MATCH')
ws3.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [8,20,36,12,20,14,24,24,22]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells('A1:I1')
hdr(ws3, 1, 1, 'Q02 — Sentence-Scene Match | Story Grammar: Initiating Event', C_HEADER, size=11)
ws3.row_dimensions[1].height = 26

ws3.merge_cells('A2:I2')
cell(ws3, 2, 1, 'Question: Read the sentence below. Which scene image does this sentence come from?', C_YELLOW, bold=True)
ws3.merge_cells('A3:I3')
cell(ws3, 3, 1, 'Target Sentence (SC02_ST01_N): "But one morning, he wakes up gray."', C_YELLOW)

q2_h = ['Option','Scene Image','Scene Description','Correct?','Score','Story Grammar Role','Distractor Rationale','Weakness Signal','LRS sg_element']
for i, h in enumerate(q2_h, 1):
    hdr(ws3, 5, i, h, C_SUB)
ws3.row_dimensions[5].height = 30

q2_opts = [
    ['A','OG0021_SC02_I','Milo wakes up gray, looks sad','YES',100,'Initiating Event','—','—','initiating_event'],
    ['B','OG0021_SC03_I','Milo in forest, meets butterfly','NO',20,'Attempt','Forest scene; student conflates gray mood with search','Understands problem led to action; misidentifies exact scene','attempt'],
    ['C','OG0021_SC06_I','Milo sits sad by blue pond','NO',30,'Consequence/Reaction','Also sad/gray in tone — emotionally similar','Focuses on emotion not specific event; shallow literal comprehension','consequence'],
    ['D','OG0021_SC09_I','Colors return — Milo glowing','NO',0,'Resolution','Opposite of gray; confuses color change direction','No story direction understanding; likely guessing','resolution'],
]
for i, r in enumerate(q2_opts, 6):
    score_bg = C_CORRECT if r[3]=='YES' else (C_PARTIAL if r[4]>=30 else ('FFEECC' if r[4]>0 else 'FFCCCC'))
    for j, v in enumerate(r, 1):
        bg = score_bg if j in [1,4,5] else (C_YELLOW if r[3]=='YES' else None)
        cell(ws3, i, j, v, bg)
    ws3.row_dimensions[i].height = 36

brd(ws3, 1, 9, 1, 9)
ws3.merge_cells('A11:I11')
hdr(ws3, 11, 1, 'LRS: verb="answered" | object="quiz_OG0021_Q02_sent_match" | result.sg_element="initiating_event" | result.option_selected=<A/B/C/D> | result.score_raw=<n>', C_GREEN)
ws3.row_dimensions[11].height = 22
brd(ws3, 11, 11, 1, 9)

# ─── SHEET 4: Q03_UNSCRAMBLE ─────────────────────────────────────────────
ws4 = wb.create_sheet('Q03_UNSCRAMBLE')
ws4.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [6,18,18,14,14,14,18,22,24]):
    ws4.column_dimensions[col].width = w

ws4.merge_cells('A1:I1')
hdr(ws4, 1, 1, 'Q03 — Sentence Unscramble | Story Grammar: Attempt', C_HEADER, size=11)
ws4.row_dimensions[1].height = 26

ws4.merge_cells('A2:I2')
cell(ws4, 2, 1, 'Question: The words below are mixed up. Drag them into the correct order to make a sentence.', C_YELLOW, bold=True)
ws4.merge_cells('A3:I3')
cell(ws4, 3, 1, 'Source: SC04_ST02_N  |  Correct: "My color helps me fly."', C_YELLOW)
ws4.merge_cells('A4:I4')
cell(ws4, 4, 1, 'Scrambled tokens presented to student: [ helps ] [ fly. ] [ color ] [ My ] [ me ]', C_ACCENT)
for r in [2,3,4]:
    ws4.row_dimensions[r].height = 22

ws4.merge_cells('A6:I6')
hdr(ws4, 6, 1, 'SECTION A — Per-Word Position Weight', C_SUB, size=10)
ws4.row_dimensions[6].height = 22

for i, h in enumerate(['Word Token','Correct Pos','Grammatical Role','Word Weight','Weight Rationale','Max Points','Cumulative'], 1):
    hdr(ws4, 7, i, h, C_SUB)

word_data = [
    ['My',     1, 'Possessive determiner (opener)', 2.0, 'Marks ownership/perspective; wrong placement = syntax confusion', '=D8/SUM($D$8:$D$12)*100', '=F8'],
    ['color',  2, 'Subject noun',                   1.5, 'Content word; confusion with "me" = noun/pronoun error',          '=D9/SUM($D$8:$D$12)*100', '=F8+F9'],
    ['helps',  3, 'Main verb',                       2.0, 'Verb position critical for SVO structure',                        '=D10/SUM($D$8:$D$12)*100','=F8+F9+F10'],
    ['me',     4, 'Object pronoun',                  1.5, 'Swapping with "color" = common noun/pronoun confusion',           '=D11/SUM($D$8:$D$12)*100','=F8+F9+F10+F11'],
    ['fly.',   5, 'Infinitive complement (closer)',  2.0, 'Closing content word; missing = incomplete predicate',            '=D12/SUM($D$8:$D$12)*100','=SUM(F8:F12)'],
]
for i, r in enumerate(word_data, 8):
    for j, v in enumerate(r, 1):
        c = ws4.cell(i, j, value=v)
        c.fill = PatternFill('solid', fgColor=C_YELLOW if j in [1,3] else C_GREEN if j == 6 else None) if j in [1,3,6] else PatternFill()
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal='center' if j not in [1,3,5] else 'left', vertical='center', wrap_text=True)
    ws4.row_dimensions[i].height = 30

for col, val in [(1,'TOTAL'), (4,'=SUM(D8:D12)'), (6,'=SUM(F8:F12)'), (7,'100 check')]:
    c = ws4.cell(13, col, value=val)
    c.fill = PatternFill('solid', fgColor=C_ACCENT)
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
brd(ws4, 6, 13, 1, 7)

ws4.merge_cells('A15:I15')
hdr(ws4, 15, 1, 'SECTION B — Partial Score Examples', C_SUB, size=10)
ws4.row_dimensions[15].height = 22

for i, h in enumerate(['Student Answer','Correct Positions','Score','Weakness Signal'], 1):
    hdr(ws4, 16, i, h, C_ACCENT, '000000')

ex4 = [
    ['My color helps me fly.', '5/5', 100, '—'],
    ['My color helps fly. me', '4/5 (me/fly swapped)', 73, 'SVO complement ordering'],
    ['color My helps me fly.', '3/5 (My/color swapped)', 60, 'Possessive determiner syntax'],
    ['helps My color me fly.', '2/5 (verb fronted)', 35, 'Basic SVO structure'],
    ['fly. me color helps My', '0/5', 0, 'No syntactic understanding'],
]
for i, r in enumerate(ex4, 17):
    score_bg = C_CORRECT if r[2]==100 else (C_PARTIAL if r[2]>=50 else 'FFCCCC')
    for j, v in enumerate(r, 1):
        bg = score_bg if j == 3 else None
        cell(ws4, i, j, v, bg)
    ws4.row_dimensions[i].height = 22

brd(ws4, 15, 21, 1, 4)
ws4.merge_cells('A23:I23')
hdr(ws4, 23, 1, 'LRS: verb="answered" | object="quiz_OG0021_Q03_unscramble" | result.sg_element="attempt" | result.word_order_submitted=[] | result.score_raw=<n> | result.words_correct=<n>/5', C_GREEN)
ws4.row_dimensions[23].height = 22
brd(ws4, 23, 23, 1, 9)

# ─── SHEET 5: Q04_MAIN_IDEA ──────────────────────────────────────────────
ws5 = wb.create_sheet('Q04_MAIN_IDEA')
ws5.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [8,36,12,20,20,24,24,22,20]):
    ws5.column_dimensions[col].width = w

ws5.merge_cells('A1:I1')
hdr(ws5, 1, 1, 'Q04 — Main Idea MCQ | Story Grammar: Theme / Resolution', C_HEADER, size=11)
ws5.row_dimensions[1].height = 26

ws5.merge_cells('A2:I2')
cell(ws5, 2, 1, 'Question: What is the main message of this story?', C_YELLOW, bold=True)

q_h = ['Option','Option Text','Correct?','Score','Story Grammar Role','Distractor Type','Distractor Rationale','Weakness Signal','LRS sg_element']
for i, h in enumerate(q_h, 1):
    hdr(ws5, 4, i, h, C_SUB)
ws5.row_dimensions[4].height = 30

q4_opts = [
    ['A','Chameleons can change their colors.','NO',10,'Setting detail','Literal distractor','Takes animal fact instead of theme','Surface reader; cannot distinguish theme from setting','setting'],
    ['B','Asking for help is important.','NO',30,'Attempt misread','Partial theme','Recognizes attempts but misses internal resolution','Focuses on action not internal discovery; mid-level comprehension','attempt'],
    ['C','Your true colors come from inside you.','YES',100,'Theme / Resolution','Correct','—','—','theme'],
    ['D','Colors are found in nature — butterflies and flowers.','NO',20,'Detail distractor','Plot detail','Remembers specific scenes; cannot extract abstract meaning','Episodic memory strong; thematic inference weak','attempt'],
]
for i, r in enumerate(q4_opts, 5):
    score_bg = C_CORRECT if r[2]=='YES' else (C_PARTIAL if r[3]>=30 else ('FFEECC' if r[3]>0 else 'FFCCCC'))
    for j, v in enumerate(r, 1):
        bg = score_bg if j in [1,3,4] else None
        cell(ws5, i, j, v, bg)
    ws5.row_dimensions[i].height = 36

brd(ws5, 1, 8, 1, 9)
ws5.merge_cells('A10:I10')
hdr(ws5, 10, 1, 'LRS: verb="answered" | object="quiz_OG0021_Q04_main_idea" | result.sg_element="theme" | result.option_selected=<A/B/C/D> | result.score_raw=<n> | result.semantic_distance=<0.0-1.0>', C_GREEN)
ws5.row_dimensions[10].height = 22
brd(ws5, 10, 10, 1, 9)

# ─── SHEET 6: Q05_CHAR_GOAL ──────────────────────────────────────────────
ws6 = wb.create_sheet('Q05_CHAR_GOAL')
ws6.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [8,36,12,20,20,24,24,22,20]):
    ws6.column_dimensions[col].width = w

ws6.merge_cells('A1:I1')
hdr(ws6, 1, 1, 'Q05 — Character Goal MCQ | Story Grammar: Initiating Event / Goal', C_HEADER, size=11)
ws6.row_dimensions[1].height = 26

ws6.merge_cells('A2:I2')
cell(ws6, 2, 1, 'Question: What does Milo want to find at the beginning of the story?', C_YELLOW, bold=True)

for i, h in enumerate(q_h, 1):
    hdr(ws6, 4, i, h, C_SUB)
ws6.row_dimensions[4].height = 30

q5_opts = [
    ['A','His lost color','YES',100,'Initiating Event / Goal','Correct','—','—','initiating_event'],
    ['B','A new friend in the forest','NO',25,'Attempt misread','Setting distractor','Remembers forest scene; confuses goal with setting','Cannot extract protagonist goal from setting details','setting'],
    ['C','Food for butterflies and flowers','NO',0,'Unrelated detail','Wild distractor','No narrative connection','Random guess or no comprehension','—'],
    ['D','A place to sit and rest','NO',15,'Consequence misread','Partial event','Recalls pond scene (sitting) but misses emotional cause','Story comprehension at surface action level only','consequence'],
]
for i, r in enumerate(q5_opts, 5):
    score_bg = C_CORRECT if r[2]=='YES' else (C_PARTIAL if r[3]>=25 else ('FFEECC' if r[3]>0 else 'FFCCCC'))
    for j, v in enumerate(r, 1):
        bg = score_bg if j in [1,3,4] else None
        cell(ws6, i, j, v, bg)
    ws6.row_dimensions[i].height = 36

brd(ws6, 1, 8, 1, 9)
ws6.merge_cells('A10:I10')
hdr(ws6, 10, 1, 'LRS: verb="answered" | object="quiz_OG0021_Q05_char_goal" | result.sg_element="initiating_event" | result.option_selected=<A/B/C/D> | result.score_raw=<n>', C_GREEN)
ws6.row_dimensions[10].height = 22
brd(ws6, 10, 10, 1, 9)

# ─── SHEET 7: Q06_EMOTION ────────────────────────────────────────────────
ws7 = wb.create_sheet('Q06_EMOTION')
ws7.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [8,36,12,20,20,24,24,22,20]):
    ws7.column_dimensions[col].width = w

ws7.merge_cells('A1:I1')
hdr(ws7, 1, 1, 'Q06 — Emotion Identification MCQ | Story Grammar: Reaction / Internal Response', C_HEADER, size=11)
ws7.row_dimensions[1].height = 26

ws7.merge_cells('A2:I2')
cell(ws7, 2, 1, 'Question: Look at scene SC06. How does Milo feel in this scene?', C_YELLOW, bold=True)
ws7.merge_cells('A3:I3')
cell(ws7, 3, 1, 'Source: SC06_Emotion = Neutral, Sad  |  Scene text (SC06_ST01-ST06_N): "Milo is sad. Everyone has their own color. He sits by a blue pond. He cries."', C_ACCENT)
ws7.row_dimensions[3].height = 22

for i, h in enumerate(q_h, 1):
    hdr(ws7, 5, i, h, C_SUB)
ws7.row_dimensions[5].height = 30

q6_opts = [
    ['A','Happy and excited','NO',0,'Internal Response mismatch','Opposite emotion','Confuses resolution emotion (SC09) with this scene','Cannot map emotions to specific moments; temporal confusion','internal_response'],
    ['B','Sad and disappointed','YES',100,'Reaction / Internal Response','Correct','SC06_Emotion=Sad confirmed in source data','—','reaction'],
    ['C','Angry and frustrated','NO',40,'Partial emotion read','Adjacent negative emotion','Recognizes negative valence but mislabels intensity','Emotion at valence level only; cannot distinguish sad vs angry','reaction'],
    ['D','Curious and hopeful','NO',20,'Anticipatory read','Forward-looking distractor','Projects pond discovery onto this moment','Skips current state; reads ahead narratively','internal_response'],
]
for i, r in enumerate(q6_opts, 6):
    score_bg = C_CORRECT if r[2]=='YES' else (C_PARTIAL if r[3]>=40 else ('FFEECC' if r[3]>0 else 'FFCCCC'))
    for j, v in enumerate(r, 1):
        bg = score_bg if j in [1,3,4] else None
        cell(ws7, i, j, v, bg)
    ws7.row_dimensions[i].height = 40

brd(ws7, 1, 9, 1, 9)
ws7.merge_cells('A11:I11')
hdr(ws7, 11, 1, 'LRS: verb="answered" | object="quiz_OG0021_Q06_emotion" | result.sg_element="reaction" | result.option_selected=<A/B/C/D> | result.score_raw=<n> | result.semantic_distance=<0.0-1.0>', C_GREEN)
ws7.row_dimensions[11].height = 22
brd(ws7, 11, 11, 1, 9)

# ─── SHEET 8: SCORING_RULES ──────────────────────────────────────────────
ws8 = wb.create_sheet('SCORING_RULES')
ws8.sheet_view.showGridLines = False
for col, w in zip('ABCDEFG', [10,24,30,18,18,28,28]):
    ws8.column_dimensions[col].width = w

ws8.merge_cells('A1:G1')
hdr(ws8, 1, 1, 'Global Scoring Rules — OG0021 Reading Quiz', C_HEADER, size=11)
ws8.row_dimensions[1].height = 26

ws8.merge_cells('A3:G3')
hdr(ws8, 3, 1, 'SECTION A — Score Band Interpretation', C_SUB, size=10)
ws8.row_dimensions[3].height = 22

for i, h in enumerate(['Band','Score Range','Color','LRS Risk Signal','SG Insight','MRI Profile Impact','Recommendation'], 1):
    hdr(ws8, 4, i, h, C_ACCENT, '000000')

bands = [
    ['Mastery','85-100','Green','None','Full SG element understood','V+/L+ in profile','Proceed to next level'],
    ['Developing','60-84','Gold','Meaning Signal (mild)','Partial narrative understanding','V0 — maintain level','Review target scene'],
    ['Emerging','30-59','Orange','Meaning Signal (moderate)','Surface/literal only','V- — flag for booster','Assign Booster Content'],
    ['Beginning','0-29','Red','Meaning Signal (strong)','Fundamental gap; guessing','V-- — priority intervention','Comprehension Booster + re-read'],
]
for i, r in enumerate(bands, 5):
    bgs = [None, C_CORRECT if i==5 else (C_YELLOW if i==6 else (C_ORANGE if i==7 else 'FFCCCC')), None, None, None, None, None]
    for j, v in enumerate(r, 1):
        cell(ws8, i, j, v, bgs[j-1])
    ws8.row_dimensions[i].height = 28

brd(ws8, 3, 8, 1, 7)

ws8.merge_cells('A10:G10')
hdr(ws8, 10, 1, 'SECTION B — Weighting Rationale by Story Grammar Element', C_SUB, size=10)
ws8.row_dimensions[10].height = 22

for i, h in enumerate(['SG Element','Quiz Coverage','Why Weighted','LRS Engine','Weakness Signal if Missed'], 1):
    hdr(ws8, 11, i, h, C_ACCENT, '000000')

wr_data = [
    ['Setting','Q01 (SC01, pos 1)','Entry anchor — gate to all other elements','Vocab & Comprehension Engine','Literal comprehension dependency'],
    ['Initiating Event / Goal','Q01 (SC03), Q02, Q05','Core problem; misidentifying = no narrative schema','Vocab & Comprehension Engine','Inferencing weakness'],
    ['Attempt','Q01 (SC05), Q03, Q04(partial)','Repetitive structure — confusion = sequential processing gap','Vocab & Comprehension Engine','Semantic Decision Stability'],
    ['Consequence / Reaction','Q01 (SC06), Q06','Emotional scene — confusion = emotion inference gap','Expression Engine','Emotion inference (Expression Signal)'],
    ['Resolution','Q01 (SC10, pos 5), Q04','Exit anchor — without this, theme extraction fails','Vocab & Comprehension Engine','Inferencing / Story Transfer weakness'],
    ['Theme','Q04','Highest-order comprehension — abstraction beyond literal','V&C Engine + Expression Engine','Limited reasoning (Expression Signal)'],
    ['Internal Response (Emotion)','Q06','Moment-level emotion mapping','Expression Engine','Perspective taking weakness'],
]
for i, r in enumerate(wr_data, 12):
    for j, v in enumerate(r, 1):
        cell(ws8, i, j, v, C_YELLOW if j==1 else None)
    ws8.row_dimensions[i].height = 26

brd(ws8, 10, 18, 1, 5)

ws8.merge_cells('A20:G20')
hdr(ws8, 20, 1, 'SECTION C — MRI Profile Contribution (per Slide 23 LRS Policy)', C_SUB, size=10)
ws8.row_dimensions[20].height = 22

ws8.merge_cells('A21:G21')
cell(ws8, 21, 1, 'Quiz scores feed into Vocabulary & Comprehension Engine (V-score) and partially Expression Engine (L-score).', C_ACCENT)
ws8.merge_cells('A22:G22')
cell(ws8, 22, 1, 'Parent Report (3-tier): Level (e.g. Lv4 — Vocabulary Retrieval) | Detail Profile: P3/V1/L2 | Risk Points: Inferencing weakness / Emotion identification gap', C_GREEN)
for r in [21, 22]:
    ws8.row_dimensions[r].height = 28
brd(ws8, 20, 22, 1, 7)

# ─── SHEET 9: LRS_MAPPING ────────────────────────────────────────────────
ws9 = wb.create_sheet('LRS_MAPPING')
ws9.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGH', [8,20,28,28,16,16,28,30]):
    ws9.column_dimensions[col].width = w

ws9.merge_cells('A1:H1')
hdr(ws9, 1, 1, 'LRS xAPI Mapping — OG0021 Quiz Statements', C_HEADER, size=11)
ws9.row_dimensions[1].height = 26

for i, h in enumerate(['Q_ID','xAPI Verb','xAPI Object','result.sg_element','result.score_raw','result.correct','Extra Fields','Risk Signal'], 1):
    hdr(ws9, 2, i, h, C_SUB)
ws9.row_dimensions[2].height = 28

lrs_rows = [
    ['Q01','answered','quiz_OG0021_Q01_sequencing','setting,initiating_event,attempt,consequence,resolution','0-100','partial','sequence_submitted[], position_scores{}','Meaning Signal if <60'],
    ['Q02','answered','quiz_OG0021_Q02_sent_match','initiating_event','0-100','true/false','option_selected, response_latency_ms','Meaning Signal if option!=A'],
    ['Q03','answered','quiz_OG0021_Q03_unscramble','attempt','0-100','partial','word_order_submitted[], words_correct/5','Meaning Signal if <60'],
    ['Q04','answered','quiz_OG0021_Q04_main_idea','theme','0-100','true/false','option_selected, semantic_distance','Expression Signal if <30 (limited reasoning)'],
    ['Q05','answered','quiz_OG0021_Q05_char_goal','initiating_event','0-100','true/false','option_selected, response_latency_ms','Meaning Signal if option!=A'],
    ['Q06','answered','quiz_OG0021_Q06_emotion','reaction','0-100','true/false','option_selected, semantic_distance','Expression Signal if <40 (emotion inference)'],
]
for i, r in enumerate(lrs_rows, 3):
    for j, v in enumerate(r, 1):
        cell(ws9, i, j, v, C_YELLOW if j==1 else None)
    ws9.row_dimensions[i].height = 28

brd(ws9, 1, 8, 1, 8)

ws9.merge_cells('A10:H10')
cell(ws9, 10, 1, 'Single LRS storage (xAPI). V&C Engine reads sg_element + score_raw for Q01-Q05. Expression Engine reads Q04/Q06 semantic_distance. MRI report: V-level (Q01-Q05) + L-level (Q04, Q06).', C_GREEN)
ws9.row_dimensions[10].height = 30
brd(ws9, 10, 10, 1, 8)

wb.save(r'C:\Users\bonni\Desktop\ISM\Content\Quiz\OG0021_ReadingQuiz.xlsx')
print('XLSX saved successfully.')
